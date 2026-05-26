# -*- coding: utf-8 -*-
"""
Created on Mon May 19 17:26:04 2025

@author: Engº Lutonda Tomalela

github: https://github.com/lutondatomalela/WallsEC2
"""


import io
import json
import math
import re
import webbrowser
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_TITLE = "WallsEC2 - Paredes de Betão Armado (EC2)"
APP_NAME = "WallsEC2"
APP_VERSION = "v1.0"
APP_AUTHOR = "Eng.º Lutonda Tomalela"
APP_SUBJECT = "Dimensionamento e verificação de paredes de betão armado segundo o Eurocódigo 2"
APP_KEYWORDS = "WallsEC2, Eurocódigo 2, EC2, NP EN 1992-1-1, paredes de betão armado, armaduras, esforço transverso, fendilhação"
APP_CATEGORY = "Structural Engineering / Reinforced Concrete Design"
APP_XLSX_DESCRIPTION = "Workbook de cálculo com dados de entrada, validação, resultados por painel, armaduras, fendilhação, corte e diagnóstico."
GITHUB_URL = "https://github.com/lutondatomalela/WallsEC2"
MAX_PREVIEW_ROWS = 20000


# ============================================================
# Utilidades gerais
# ============================================================
def normalize_text(s: str) -> str:
    s = str(s).strip().lower()
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def safe_float(value, default=float("nan")):
    try:
        if pd.isna(value):
            return default
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return default
        s = s.replace("\u00a0", " ").replace(" ", "")
        # Formato PT: 1.234,56
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+,\d+", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return default


def bar_area_mm2(phi_mm: float) -> float:
    return math.pi * phi_mm * phi_mm / 4.0


def parse_strength_from_class(cls: str, default: float = 30.0) -> float:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)", str(cls))
    if m:
        return float(m.group(1).replace(",", "."))
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(cls))
    if m:
        return float(m.group(1).replace(",", "."))
    return default


def concrete_props(fck: float, alpha_cc: float = 1.0, gamma_c: float = 1.5) -> Dict[str, float]:
    fcm = fck + 8.0
    fcd = alpha_cc * fck / gamma_c
    fctm = 0.30 * fck ** (2.0 / 3.0) if fck <= 50 else 2.12 * math.log(1 + fcm / 10.0)
    return {"fck": fck, "fcm": fcm, "fcd": fcd, "fctm": fctm}


def steel_fyd(fyk: float, gamma_s: float = 1.15) -> float:
    return fyk / gamma_s


# ============================================================
# Importação / limpeza da tabela software de análise estrutural
# ============================================================
COLUMN_ALIASES = {
    "panel_node_case": [
        "panel/node/case", "panel / node / case", "panel/node", "panel node case",
        "object/node/case", "object / node / case", "element/node/case",
        "finite element/node/case", "fe/node/case", "painel/nó/caso", "painel/no/caso"
    ],
    "panel": ["panel", "painel", "object", "element", "finite element", "fe", "objeto", "elemento"],
    "node": ["node", "nó", "no"],
    "case": ["case", "caso", "combination", "combinação", "combinacao"],
    "mxx": ["mxx (knm/m)", "mxx [knm/m]", "mxx", "mx (knm/m)", "mx [knm/m]", "mx"],
    "myy": ["myy (knm/m)", "myy [knm/m]", "myy", "my (knm/m)", "my [knm/m]", "my"],
    "mxy": ["mxy (knm/m)", "mxy [knm/m]", "mxy", "mxyy"],
    "qxx": ["qxx (kn/m)", "qxx [kn/m]", "qxx", "qx (kn/m)", "qx [kn/m]", "qx", "vxx", "vx"],
    "qyy": ["qyy (kn/m)", "qyy [kn/m]", "qyy", "qy (kn/m)", "qy [kn/m]", "qy", "vyy", "vy"],
    "nxx": ["nxx (kn/m)", "nxx [kn/m]", "nxx", "nx (kn/m)", "nx [kn/m]", "nx"],
    "nyy": ["nyy (kn/m)", "nyy [kn/m]", "nyy", "ny (kn/m)", "ny [kn/m]", "ny"],
    "nxy": ["nxy (kn/m)", "nxy [kn/m]", "nxy"],
}


def parse_pasted_table(text: str) -> pd.DataFrame:
    text = text.strip()
    if not text:
        return pd.DataFrame()

    # Excel/tabela de cálculo copiado normalmente vem separado por tabs, ponto e vírgula ou barras verticais.
    for sep in ("\t", ";", "|"):
        try:
            df = pd.read_csv(io.StringIO(text), sep=sep, engine="python", dtype=str)
            if len(df.columns) > 1:
                return df
        except Exception:
            pass

    lines = [ln.rstrip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return pd.DataFrame()

    # Leitura robusta para tabelas copiadas do software de análise estrutural com espaços simples, por exemplo:
    # 43/ 49/ 302 (CQC) -2,17 -0,13 0,83 43 5,59 3,69
    robot_rows = []
    row_re = re.compile(
        r"^\s*(?P<panel>\S+)\s*/\s*(?P<node>\S+)\s*/\s*(?P<case>\d+)\s*"
        r"(?P<casetype>\([^)]*\))?\s+(?P<rest>.*)$"
    )
    num_re = re.compile(r"[-+]?\d+(?:[\.,]\d+)?")
    for ln in lines[1:]:
        m = row_re.match(ln)
        if not m:
            robot_rows = []
            break
        nums = num_re.findall(m.group("rest"))
        # Esperado: MXX, MYY, MXY, Panel, QXX, QYY. Em algumas tabelas o Panel repetido pode faltar.
        if len(nums) >= 6:
            mxx, myy, mxy, panel_rep, qxx, qyy = nums[:6]
        elif len(nums) >= 5:
            mxx, myy, mxy, qxx, qyy = nums[:5]
            panel_rep = m.group("panel")
        else:
            robot_rows = []
            break
        case_txt = m.group("case")
        if m.group("casetype"):
            case_txt = f"{case_txt} {m.group('casetype').strip()}"
        robot_rows.append({
            "panel_node_case": f"{m.group('panel')}/{m.group('node')}/{case_txt}",
            "panel": panel_rep,
            "node": m.group("node"),
            "case": case_txt,
            "mxx": mxx,
            "myy": myy,
            "mxy": mxy,
            "qxx": qxx,
            "qyy": qyy,
        })
    if robot_rows:
        return pd.DataFrame(robot_rows)

    # Fallback para espaços múltiplos.
    rows = [re.split(r"\s{2,}", ln.strip()) for ln in lines]
    header = rows[0]
    body = rows[1:]
    width = len(header)
    body = [r[:width] + [""] * max(0, width - len(r)) for r in body]
    return pd.DataFrame(body, columns=header)


def rename_known_columns(df: pd.DataFrame) -> pd.DataFrame:
    norm_to_original = {normalize_text(c): c for c in df.columns}
    rename_map = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm_to_original:
                rename_map[norm_to_original[alias]] = target
                break
    # fallback: remove units/brackets and retry, useful for software de análise estrutural tables with headers like MXX [kNm/m].
    for original in df.columns:
        if original in rename_map:
            continue
        simplified = normalize_text(re.sub(r"[\[\(].*?[\]\)]", "", str(original)))
        for target, aliases in COLUMN_ALIASES.items():
            if target in rename_map.values():
                continue
            if simplified in [normalize_text(re.sub(r"[\[\(].*?[\]\)]", "", a)) for a in aliases]:
                rename_map[original] = target
                break
    return df.rename(columns=rename_map).copy()


def split_panel_node_case(s: str) -> Tuple[str, str, str]:
    parts = [p.strip() for p in str(s).split("/")]
    panel = parts[0] if len(parts) > 0 else ""
    node = parts[1] if len(parts) > 1 else ""
    case = parts[2] if len(parts) > 2 else ""
    return panel, node, case


def canonical_case_id(value: object) -> str:
    """Extrai o número do caso/combinação, ignorando textos como (C), (CQC)."""
    txt = normalize_text(str(value))
    txt = txt.replace("case", "").replace("caso", "").replace("comb.", "").replace("comb", "")
    m = re.search(r"\d+", txt)
    return m.group(0) if m else txt.strip()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = rename_known_columns(df)
    df["__row_order"] = range(len(df))

    if "panel_node_case" in df.columns:
        vals = df["panel_node_case"].map(split_panel_node_case)
        df["panel_from_id"] = vals.map(lambda x: x[0])
        if "node" not in df.columns:
            df["node"] = vals.map(lambda x: x[1])
        if "case" not in df.columns:
            df["case"] = vals.map(lambda x: x[2])
    else:
        df["panel_from_id"] = ""
        if "node" not in df.columns:
            df["node"] = ""
        if "case" not in df.columns:
            df["case"] = ""

    if "panel" not in df.columns:
        df["panel"] = df["panel_from_id"]
    df["panel"] = df["panel"].astype(str).replace("nan", "")
    df.loc[df["panel"].str.strip() == "", "panel"] = df["panel_from_id"]
    if "case" in df.columns:
        df["case_id"] = df["case"].map(canonical_case_id)

    for c in ["mxx", "myy", "mxy", "qxx", "qyy", "nxx", "nyy", "nxy"]:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = df[c].map(lambda v: safe_float(v, 0.0))

    return df


def reduce_to_governing_cases(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    work = df.copy()
    work["_mmax"] = work[["mxx", "myy", "mxy"]].abs().max(axis=1)
    work["_qmax"] = work[["qxx", "qyy"]].abs().max(axis=1)
    work["_score"] = work["_mmax"] + 0.15 * work["_qmax"]
    idx = set()
    for _, grp in work.groupby(["panel"], dropna=False):
        if grp.empty:
            continue
        idx.add(grp["_score"].idxmax())
        idx.add(grp["mxx"].abs().idxmax())
        idx.add(grp["myy"].abs().idxmax())
        idx.add(grp["mxy"].abs().idxmax())
        idx.add(grp["qxx"].abs().idxmax())
        idx.add(grp["qyy"].abs().idxmax())
    out = work.loc[sorted(idx)].copy().sort_values(["panel", "case", "node", "__row_order"])
    return out.drop(columns=["_mmax", "_qmax", "_score"], errors="ignore").reset_index(drop=True)


# ============================================================
# Dimensionamento EC2 simplificado para parede como elemento de placa
# ============================================================
@dataclass
class RebarSolution:
    phi: float
    spacing: float
    as_prov: float
    text_override: str = ""
    base_text: str = ""
    add_text: str = ""
    as_base: float = 0.0
    as_add: float = 0.0
    optimization_note: str = ""
    n_iterations: int = 0

    @property
    def text(self) -> str:
        if self.text_override:
            return self.text_override
        return f"Ø{int(self.phi)}//{int(self.spacing)}"


def moment_unit_factor_to_kNm_per_m(unit: str) -> float:
    u = normalize_text(unit)
    if "nmm/mm" in u:
        return 0.001
    if "nm/m" in u and "kn" not in u:
        return 0.001
    if "knmm/m" in u:
        return 0.001
    if "knm/m" in u or "kn.m/m" in u or "kn·m/m" in u:
        return 1.0
    return 1.0


def shear_unit_factor_to_kN_per_m(unit: str) -> float:
    u = normalize_text(unit)
    if "n/mm" in u:
        return 1.0
    if "n/m" in u and "kn" not in u:
        return 0.001
    if "kn/m" in u:
        return 1.0
    return 1.0


def status_rank(status: str) -> int:
    order = {"OK": 0, "OK*": 1, "Verificar": 2, "Dados insuficientes": 3, "Não conforme": 4}
    return order.get(str(status), 2)


class WallDesigner:
    def __init__(
        self,
        thickness_mm: float,
        cover_mm: float,
        concrete_class: str,
        fyk: float,
        gamma_c: float = 1.5,
        gamma_s: float = 1.15,
        alpha_cc: float = 1.0,
        local_y_is_vertical: bool = True,
        aggregate_mm: float = 20.0,
        crack_spacing_limit_mm: Optional[float] = None,
        wood_armer_method: str = "Conservativo |MXY|",
        swap_local_axes: bool = False,
        crack_check_enabled: bool = True,
        wk_check_enabled: bool = False,
        qp_case: str = "",
        moment_unit: str = "kNm/m",
        shear_unit: str = "kN/m",
        combo_type: str = "ELU",
        wmax_mm: float = 0.30,
        phi_min_mm: float = 8.0,
        phi_max_mm: float = 16.0,
        optimize_rebar: bool = False,
        rebar_strategy: str = "Base + reforços",
        base_vertical_phi: float = 10.0,
        base_vertical_spacing: float = 200.0,
        base_horizontal_phi: float = 8.0,
        base_horizontal_spacing: float = 200.0,
    ):
        self.t = thickness_mm
        self.cover = cover_mm
        self.fck = parse_strength_from_class(concrete_class, 30.0)
        self.cp = concrete_props(self.fck, alpha_cc=alpha_cc, gamma_c=gamma_c)
        self.fyd = steel_fyd(fyk, gamma_s)
        self.fyk = fyk
        self.local_y_is_vertical = local_y_is_vertical
        self.aggregate_mm = aggregate_mm
        self.crack_spacing_limit_mm = crack_spacing_limit_mm
        self.wood_armer_method = wood_armer_method
        self.swap_local_axes = swap_local_axes
        self.crack_check_enabled = crack_check_enabled
        self.wk_check_enabled = wk_check_enabled
        self.qp_case = str(qp_case).strip()
        self.moment_unit = moment_unit
        self.shear_unit = shear_unit
        self.moment_factor = moment_unit_factor_to_kNm_per_m(moment_unit)
        self.shear_factor = shear_unit_factor_to_kN_per_m(shear_unit)
        self.combo_type = combo_type
        self.wmax_mm = wmax_mm
        base_phi = [8.0, 10.0, 12.0, 16.0, 20.0, 25.0]
        self.phi_candidates = [p for p in base_phi if phi_min_mm <= p <= phi_max_mm] or [8.0, 10.0, 12.0, 16.0]
        self.spacing_candidates = [75.0, 100.0, 125.0, 150.0, 175.0, 200.0, 225.0, 250.0, 300.0, 350.0, 400.0]
        self.optimize_rebar = bool(optimize_rebar)
        self.rebar_strategy = str(rebar_strategy)
        self.base_vertical_phi = float(base_vertical_phi)
        self.base_vertical_spacing = float(base_vertical_spacing)
        self.base_horizontal_phi = float(base_horizontal_phi)
        self.base_horizontal_spacing = float(base_horizontal_spacing)

    @property
    def ac_per_m(self) -> float:
        return 1000.0 * self.t

    def min_wall_reinf_total(self) -> Tuple[float, float, float]:
        """Retorna mínimos totais por metro: vertical, horizontal e máximo vertical."""
        ac = self.ac_per_m
        # EC2 9.6.2: As,vmin recomendado = 0,002 Ac; As,vmax recomendado = 0,04 Ac.
        asv_min = 0.002 * ac
        asv_max = 0.04 * ac
        # EC2 9.6.3: As,hmin >= max(25% da armadura vertical; 0,001 Ac).
        ash_min = max(0.25 * asv_min, 0.001 * ac)
        return asv_min, ash_min, asv_max

    def slab_flexural_min_per_face(self, phi: float = 12.0) -> float:
        """Mínimo de flexão tipo laje/faixa de 1 m: EC2 9.3.1.1."""
        d = self.effective_depth(phi)
        return max(0.26 * self.cp["fctm"] / self.fyk * 1000.0 * d, 0.0013 * 1000.0 * d)

    def max_spacing_for_direction(self, is_vertical_direction: bool) -> float:
        # EC2 9.6.2: varões verticais <= min(3t, 400 mm).
        # EC2 9.6.3: varões horizontais <= 400 mm.
        base = min(3.0 * self.t, 400.0) if is_vertical_direction else 400.0
        if self.crack_spacing_limit_mm and self.crack_spacing_limit_mm > 0:
            base = min(base, self.crack_spacing_limit_mm)
        return base

    def effective_depth(self, phi: float = 12.0) -> float:
        # Armadura numa face, flexão transversal da parede/laje.
        return max(self.t - self.cover - phi / 2.0, 1.0)

    def flexural_as_required(self, m_ed_kNm_per_m: float, phi_assumed: float = 12.0) -> float:
        """As requerido por metro de largura, mm2/m, para flexão simples de faixa de 1 m."""
        m = abs(m_ed_kNm_per_m) * 1e6  # Nmm/m
        if m <= 1e-9:
            return 0.0
        d = self.effective_depth(phi_assumed)
        z = 0.90 * d
        return m / (0.87 * self.fyd * z)

    def mrd_kNm_per_m(self, as_mm2_per_m: float, phi: float) -> float:
        d = self.effective_depth(phi)
        fcd = self.cp["fcd"]
        # bloco retangular simplificado EC2: C = 0.8 x b fcd, b=1000 mm
        x = as_mm2_per_m * self.fyd / (0.8 * 1000.0 * fcd) if fcd > 0 else 1e9
        z = d - 0.4 * x
        if x > 0.45 * d:
            z = max(0.75 * d, z)  # mantém valor conservativo razoável e assinala via utilização se necessário
        return as_mm2_per_m * self.fyd * z / 1e6

    def choose_rebar(self, as_req_per_face: float, max_spacing: float = 400.0) -> RebarSolution:
        best = None
        for phi in self.phi_candidates:
            for s in self.spacing_candidates:
                if s > max_spacing + 1e-9:
                    continue
                clear = s - phi
                min_clear = max(phi, self.aggregate_mm + 5.0, 20.0)
                if clear < min_clear:
                    continue
                as_prov = bar_area_mm2(phi) * 1000.0 / s
                if as_prov + 1e-9 >= as_req_per_face:
                    cand = RebarSolution(phi=phi, spacing=s, as_prov=as_prov)
                    excess = cand.as_prov / max(as_req_per_face, 1.0)
                    # Engenharia prática: evita espaçamentos <100 mm, evita Ø20/Ø25 quando não necessários,
                    # e minimiza excesso de aço sem penalizar demasiado soluções correntes Ø10/Ø12.
                    penalty_spacing = 0 if cand.spacing >= 100.0 else 10
                    penalty_large_phi = max(0.0, cand.phi - 16.0) / 10.0
                    penalty_preferred = 0 if cand.phi in (8.0, 10.0, 12.0, 16.0) else 1
                    key = (penalty_spacing, penalty_large_phi, excess, cand.as_prov, penalty_preferred, cand.phi, -cand.spacing)
                    if best is None:
                        best = cand
                        best_key = key
                    elif key < best_key:
                        best = cand
                        best_key = key
        if best is None:
            # solução forçada: Ø25//100
            phi, s = 25.0, 100.0
            best = RebarSolution(phi, s, bar_area_mm2(phi) * 1000.0 / s)
        return best

    def rebar_from_phi_spacing(self, phi: float, spacing: float) -> RebarSolution:
        phi = float(phi)
        spacing = float(spacing)
        if phi <= 0 or spacing <= 0:
            return RebarSolution(0.0, 1e9, 0.0, text_override="-")
        return RebarSolution(phi, spacing, bar_area_mm2(phi) * 1000.0 / spacing)

    def base_solution_for_direction(self, is_vertical_direction: bool) -> RebarSolution:
        if is_vertical_direction:
            return self.rebar_from_phi_spacing(self.base_vertical_phi, self.base_vertical_spacing)
        return self.rebar_from_phi_spacing(self.base_horizontal_phi, self.base_horizontal_spacing)

    def base_solution_is_valid(self, sol: RebarSolution, max_spacing: float) -> Tuple[bool, str]:
        if sol.phi <= 0 or sol.spacing <= 0:
            return False, "armadura base não definida"
        if sol.spacing > max_spacing + 1e-9:
            return False, f"espaçamento base {sol.spacing:.0f} mm > smax {max_spacing:.0f} mm"
        clear = sol.spacing - sol.phi
        min_clear = max(sol.phi, self.aggregate_mm + 5.0, 20.0)
        if clear < min_clear:
            return False, f"distância livre base insuficiente ({clear:.0f} mm < {min_clear:.0f} mm)"
        return True, "armadura base válida"

    def choose_compatible_additional_rebar(self, deficit_as: float, base: RebarSolution, max_spacing: float) -> RebarSolution:
        """Escolhe reforço coerente com a malha base.

        Regra de projecto adoptada nesta versão:
        - o reforço adicional mantém o mesmo espaçamento da armadura base;
        - evita soluções do tipo Ø10//200 + Ø8//150, que são pouco coerentes em desenho/obra;
        - se uma única família adicional não for suficiente, admite 2 ou 3 famílias iguais, sempre com o mesmo espaçamento da base.

        Exemplo: base Ø10//200 + adicional Ø8//200. Na prática, as duas famílias podem ser intercaladas,
        resultando numa distância alternada próxima de 100 mm entre varões.
        """
        s = float(base.spacing)
        if s <= 0 or s > max_spacing + 1e-9:
            return self.choose_rebar(deficit_as, max_spacing=max_spacing)

        candidates = []
        for n_layers in (1, 2, 3):
            for phi in self.phi_candidates:
                # Distância livre aproximada entre varões intercalados da base e do reforço.
                # Se houver várias famílias adicionais, o espaçamento real fica ainda mais apertado,
                # pelo que se usa s/(n_layers+1) como aproximação conservativa.
                equivalent_spacing = s / (n_layers + 1)
                clear_between_families = equivalent_spacing - max(phi, base.phi)
                min_clear = max(phi, base.phi, self.aggregate_mm + 5.0, 20.0)
                if clear_between_families < min_clear:
                    continue

                as_add_single = bar_area_mm2(phi) * 1000.0 / s
                as_add = n_layers * as_add_single
                if as_add + 1e-9 < deficit_as:
                    continue

                if n_layers == 1:
                    text = f"Ø{int(phi)}//{int(s)}"
                else:
                    text = f"{n_layers}Ø{int(phi)}//{int(s)}"

                sol = RebarSolution(phi=phi, spacing=s, as_prov=as_add, text_override=text)
                excess = as_add / max(deficit_as, 1.0)
                penalty_layers = n_layers - 1
                penalty_phi = 0 if phi <= 12.0 else (phi - 12.0) / 10.0
                key = (penalty_layers, penalty_phi, excess, phi)
                candidates.append((key, sol))

        if candidates:
            candidates.sort(key=lambda x: x[0])
            return candidates[0][1]

        # Fallback: mantém a coerência de espaçamento, mas informa que a solução foi forçada.
        phi = max(self.phi_candidates) if self.phi_candidates else max(base.phi, 16.0)
        as_add = bar_area_mm2(phi) * 1000.0 / s
        n_layers = max(1, math.ceil(deficit_as / max(as_add, 1.0)))
        text = f"{n_layers}Ø{int(phi)}//{int(s)}" if n_layers > 1 else f"Ø{int(phi)}//{int(s)}"
        return RebarSolution(phi=phi, spacing=s, as_prov=n_layers * as_add, text_override=text)

    def optimize_rebar_solution(self, as_req: float, max_spacing: float, is_vertical_direction: bool, label: str) -> RebarSolution:
        """Adopta armadura automática. Se activado, usa base + reforço adicional local coerente."""
        full = self.choose_rebar(as_req, max_spacing=max_spacing)
        full.as_base = 0.0
        full.as_add = full.as_prov
        full.add_text = full.text
        full.optimization_note = "solução automática directa"
        full.n_iterations = 1

        if not self.optimize_rebar:
            return full

        if "substituir" in normalize_text(self.rebar_strategy):
            full.optimization_note = f"estratégia substituir base: adoptada solução uniforme directa para {label}"
            full.n_iterations = 1
            return full

        base = self.base_solution_for_direction(is_vertical_direction)
        ok_base, base_note = self.base_solution_is_valid(base, max_spacing)
        if not ok_base:
            full.optimization_note = f"base rejeitada ({base_note}); adoptada solução automática directa"
            full.base_text = base.text
            return full

        # Iteração 0: armadura base
        if base.as_prov + 1e-9 >= as_req:
            return RebarSolution(
                base.phi, base.spacing, base.as_prov,
                text_override=base.text, base_text=base.text, add_text="-",
                as_base=base.as_prov, as_add=0.0,
                optimization_note=f"base suficiente para {label}", n_iterations=0,
            )

        # Iteração 1: manter base e acrescentar reforço local com o MESMO espaçamento da base.
        # Isto evita combinações pouco coerentes, como Ø10//200 + Ø8//150.
        deficit = max(0.0, as_req - base.as_prov)
        add = self.choose_compatible_additional_rebar(deficit, base, max_spacing=max_spacing)
        final_as = base.as_prov + add.as_prov
        text = f"{base.text} + {add.text}"

        return RebarSolution(
            min(base.phi, add.phi), base.spacing, final_as,
            text_override=text, base_text=base.text, add_text=add.text,
            as_base=base.as_prov, as_add=add.as_prov,
            optimization_note=f"base insuficiente; reforço local coerente com a malha base para {label}", n_iterations=1,
        )

    def shear_vrdc_kN_per_m(self, as_l_per_m: float, d_mm: float) -> float:
        fck = self.fck
        gamma_c = 1.5
        crdc = 0.18 / gamma_c
        k = min(2.0, 1.0 + math.sqrt(200.0 / d_mm)) if d_mm > 0 else 2.0
        rho_l = min(0.02, max(0.0, as_l_per_m / (1000.0 * d_mm))) if d_mm > 0 else 0.0
        vrdc = crdc * k * (100.0 * rho_l * fck) ** (1.0 / 3.0) if rho_l > 0 else 0.0
        vmin = 0.035 * k ** 1.5 * math.sqrt(fck)
        vrdc = max(vrdc, vmin)  # MPa = N/mm2
        return vrdc * d_mm  # N/mm = kN/m

    def shear_vrdmax_kN_per_m(self, d_mm: float, cot_theta: float = 2.5) -> float:
        """Limite superior simplificado de esmagamento da biela comprimida, por metro."""
        if d_mm <= 0:
            return 0.0
        fcd = self.cp["fcd"]
        nu1 = 0.6 * (1.0 - self.fck / 250.0)
        z = 0.9 * d_mm
        tan_theta = 1.0 / cot_theta
        # N/mm = kN/m, para bw = 1000 mm
        return 1000.0 * z * nu1 * fcd / (cot_theta + tan_theta) / 1000.0

    def wood_armer_moments(self, mxx: float, myy: float, mxy: float) -> Tuple[float, float, float, float, str]:
        method = normalize_text(self.wood_armer_method)
        if "principal" in method:
            avg = 0.5 * (mxx + myy)
            rad = math.sqrt((0.5 * (mxx - myy)) ** 2 + mxy ** 2)
            m1 = avg + rad
            m2 = avg - rad
            # Distribuição conservativa por armadura ortogonal: adopta o principal máximo como referência nas duas direcções.
            mx_pos = max(0.0, m1)
            mx_neg = max(0.0, -m2)
            my_pos = max(0.0, m1)
            my_neg = max(0.0, -m2)
            return mx_pos, mx_neg, my_pos, my_neg, "Momentos principais conservativos"
        # Método base: envelope conservativo com |MXY|.
        mx_pos = max(0.0, mxx + abs(mxy))
        mx_neg = max(0.0, -mxx + abs(mxy))
        my_pos = max(0.0, myy + abs(mxy))
        my_neg = max(0.0, -myy + abs(mxy))
        return mx_pos, mx_neg, my_pos, my_neg, "Wood-Armer conservativo |MXY|"

    def crack_simplified_status(self, sol: RebarSolution) -> Tuple[str, str]:
        # Controlo sempre activo quando wk não é seleccionado.
        # Usa espaçamento efectivo quando existe reforço intercalado.
        phi_ctrl, s_eff, _ = self.rebar_crack_geometry(sol)
        if s_eff <= 200.0 and phi_ctrl <= 12.0:
            return "OK", "controlo simplificado: Ø≤12 e s_ef≤200 mm"
        if s_eff <= 250.0 and phi_ctrl <= 16.0:
            return "OK*", "controlo simplificado preliminar: Ø≤16 e s_ef≤250 mm; confirmar por ELS se condicionante"
        return "Verificar", "diâmetro/espaçamento efectivo elevado; recomenda-se verificação wk com combinação quase-permanente"

    def crack_min_as_per_face(self, sigma_s_lim: float = 500.0) -> float:
        # EC2 7.3.2 em forma simplificada para faixa de 1 m: kc*k*fct,eff*Act/sigma_s.
        # Act é aproximado como a meia espessura traccionada por face.
        kc = 0.40
        k = 1.00 if self.t <= 300.0 else 0.65
        fct_eff = self.cp["fctm"]
        act = 1000.0 * self.t / 2.0
        return kc * k * fct_eff * act / max(sigma_s_lim, 1.0)

    def _parse_rebar_text(self, txt: str) -> List[Tuple[int, float, float]]:
        """Extrai famílias de armadura do texto: Ø10//200, 2Ø12//200, etc."""
        out = []
        for part in str(txt or "").replace("phi", "Ø").split("+"):
            p = part.strip().replace(" ", "")
            if not p or p == "-":
                continue
            m = re.search(r"(?:(\d+)Ø)?(\d+(?:[.,]\d+)?)//(\d+(?:[.,]\d+)?)", p)
            if not m:
                m = re.search(r"(\d+)?\s*[Øø]\s*(\d+(?:[.,]\d+)?)\s*/{2,3}\s*(\d+(?:[.,]\d+)?)", p)
            if m:
                n = int(m.group(1)) if m.group(1) else 1
                phi = float(m.group(2).replace(",", "."))
                s = float(m.group(3).replace(",", "."))
                out.append((max(1, n), phi, s))
        return out

    def rebar_crack_geometry(self, sol: RebarSolution) -> Tuple[float, float, int]:
        """Retorna Ø de controlo, espaçamento efectivo e nº aproximado de famílias para fendilhação.

        Para base + reforços, assume reforços intercalados com o mesmo espaçamento da base:
        Ø10//200 + Ø8//200 -> s_eff ≈ 100 mm.
        """
        families = []
        if sol.base_text:
            families.extend(self._parse_rebar_text(sol.base_text))
        if sol.add_text and sol.add_text != "-":
            families.extend(self._parse_rebar_text(sol.add_text))
        if not families:
            families.extend(self._parse_rebar_text(sol.text))
        if not families:
            return sol.phi, sol.spacing, 1
        phi_ctrl = max(phi for n, phi, s in families)
        # se as famílias têm o mesmo espaçamento, usa intercalamento; caso contrário, usa o menor espaçamento.
        spacings = [s for n, phi, s in families if s > 0]
        n_fam = sum(n for n, phi, s in families)
        if spacings and max(spacings) - min(spacings) < 1e-6:
            s_eff = spacings[0] / max(1, n_fam)
        else:
            s_eff = min(spacings) if spacings else sol.spacing
        return phi_ctrl, s_eff, max(1, n_fam)

    def _interp_limit(self, sigma_s: float, table: List[Tuple[float, float]]) -> float:
        if not table:
            return 0.0
        sigma_s = max(0.0, float(sigma_s))
        table = sorted(table)
        if sigma_s <= table[0][0]:
            return table[0][1]
        for (s0, v0), (s1, v1) in zip(table[:-1], table[1:]):
            if s0 <= sigma_s <= s1:
                if abs(s1 - s0) < 1e-9:
                    return v0
                return v0 + (v1 - v0) * (sigma_s - s0) / (s1 - s0)
        return table[-1][1]

    def ec2_crack_limits(self, sigma_s: float) -> Tuple[float, float, str]:
        """Limites aproximados por tabelas EC2 7.3N/7.2N para controlo sem cálculo directo.

        As tabelas são usadas como verificação de coerência de diâmetro/espaçamento para o valor de wmax
        seleccionado. Para valores intermédios de tensão é feita interpolação linear.
        """
        w = float(self.wmax_mm)
        # Tabelas indicativas usuais do EC2 para betão armado.
        # sigma_s [MPa] -> phi_max [mm]
        phi_tables = {
            0.4: [(160, 40), (200, 32), (240, 20), (280, 16), (320, 12), (360, 10), (400, 8), (450, 6)],
            0.3: [(160, 32), (200, 25), (240, 16), (280, 12), (320, 10), (360, 8), (400, 6), (450, 5)],
            0.2: [(160, 25), (200, 16), (240, 12), (280, 8), (320, 6), (360, 5), (400, 4), (450, 0)],
        }
        # sigma_s [MPa] -> s_max [mm]
        spacing_tables = {
            0.4: [(160, 300), (200, 300), (240, 250), (280, 200), (320, 150), (360, 100), (400, 50), (450, 0)],
            0.3: [(160, 300), (200, 250), (240, 200), (280, 150), (320, 100), (360, 50), (400, 0), (450, 0)],
            0.2: [(160, 200), (200, 150), (240, 100), (280, 50), (320, 0), (360, 0), (400, 0), (450, 0)],
        }
        # selecciona a classe de wmax imediatamente inferior/igual, por segurança
        if w <= 0.20 + 1e-9:
            key = 0.2
        elif w <= 0.30 + 1e-9:
            key = 0.3
        else:
            key = 0.4
        phi_lim = self._interp_limit(sigma_s, phi_tables[key])
        s_lim = self._interp_limit(sigma_s, spacing_tables[key])
        note = f"limites EC2 indicativos para wmax≈{key:.1f} mm e σs={sigma_s:.0f} MPa"
        return phi_lim, s_lim, note

    def crack_width_estimate(self, m_ed_kNm_per_m: float, sol: RebarSolution) -> Tuple[float, float, float, str]:
        # Estimativa EC2 7.3.4 simplificada para flexão de faixa de 1 m.
        # Usa a armadura final real na face traccionada: As_base + As_reforço.
        # Deve ser usada apenas com esforços ELS.
        m = abs(m_ed_kNm_per_m) * 1e6
        if m <= 1e-9 or sol.as_prov <= 1e-9:
            return 0.0, 0.0, 0.0, "sem tracção relevante"
        Es = 200000.0
        Ecm = 22.0 * ((self.fck + 8.0) / 10.0) ** 0.3 * 1000.0
        alpha_e = Es / Ecm if Ecm > 0 else 6.0
        phi_ctrl, s_eff, _ = self.rebar_crack_geometry(sol)
        d = self.effective_depth(phi_ctrl)
        z = 0.90 * d
        sigma_s = min(0.95 * self.fyk, m / (z * sol.as_prov))
        h_ceff = min(2.5 * (self.t - d), self.t / 2.0)
        a_ceff = max(1.0, 1000.0 * h_ceff)
        rho_eff = max(sol.as_prov / a_ceff, 1e-5)
        k1 = 0.8
        k2 = 0.5
        k3 = 3.4
        k4 = 0.425
        kt = 0.4 if "quase" in normalize_text(self.combo_type) or self.wk_check_enabled else 0.6
        c = self.cover
        sr_formula = k3 * c + k1 * k2 * k4 * phi_ctrl / rho_eff
        # Não se força sr_max ao espaçamento, mas reporta-se s_eff e compara-se com o limite EC2.
        sr_max = sr_formula
        eps = max((sigma_s - kt * self.cp["fctm"] / rho_eff * (1.0 + alpha_e * rho_eff)) / Es, 0.6 * sigma_s / Es)
        wk = sr_max * eps
        note = "wk estimado com As final real da face traccionada; confirmar se os esforços são de ELS quase-permanente"
        return wk, sigma_s, sr_max, note

    def crack_check_detail(self, m_ed_kNm_per_m: float, sol: RebarSolution, face: str, direction: str) -> Dict[str, object]:
        wk, sigma_s, sr_max, note = self.crack_width_estimate(m_ed_kNm_per_m, sol)
        phi_ctrl, s_eff, n_fam = self.rebar_crack_geometry(sol)
        phi_lim, s_lim, lim_note = self.ec2_crack_limits(sigma_s)
        if wk <= self.wmax_mm + 1e-9 and phi_ctrl <= phi_lim + 1e-9 and s_eff <= s_lim + 1e-9:
            st = "OK"
            reason = f"wk={wk:.3f} mm ≤ wmax; Øctrl={phi_ctrl:.0f}≤Ølim={phi_lim:.0f}; s_eff={s_eff:.0f}≤s_lim={s_lim:.0f}"
        elif wk <= self.wmax_mm + 1e-9:
            st = "OK*"
            reason = f"wk cumpre, mas diâmetro/espaçamento excede limite simplificado: Øctrl={phi_ctrl:.0f}/Ølim={phi_lim:.0f}; s_eff={s_eff:.0f}/s_lim={s_lim:.0f}"
        else:
            st = "Verificar"
            reason = f"wk={wk:.3f} mm > wmax={self.wmax_mm:.3f} mm"
        return {
            "Face": face, "Direction": direction, "Rebar": sol.text,
            "As_total_mm2_m": sol.as_prov, "As_base_mm2_m": sol.as_base, "As_add_mm2_m": sol.as_add,
            "phi_control_mm": phi_ctrl, "s_eff_mm": s_eff, "n_families": n_fam,
            "sigma_s_MPa": sigma_s, "phi_lim_EC2_mm": phi_lim, "s_lim_EC2_mm": s_lim,
            "sr_max_mm": sr_max, "wk_mm": wk, "wmax_mm": self.wmax_mm,
            "Status": st, "Reason": reason, "Note": note + "; " + lim_note,
        }

    def optimize_solution_for_cracking(self, m_qp: float, sol: RebarSolution, as_req: float, max_spacing: float, is_vertical_direction: bool, label: str) -> RebarSolution:
        """Optimização específica de fendilhação para a combinação quase-permanente.

        Prioriza: mais varões, menor diâmetro, reforço intercalado e coerência com a malha base.
        """
        if not self.wk_check_enabled or abs(m_qp) <= 1e-9:
            return sol
        current_detail = self.crack_check_detail(m_qp, sol, label[-1:] if label else "", label[:1] if label else "")
        if current_detail["Status"] in ("OK", "OK*") and current_detail["wk_mm"] <= self.wmax_mm + 1e-9:
            return sol

        candidates: List[RebarSolution] = [sol]
        # soluções directas com espaçamentos menores
        for phi in sorted(self.phi_candidates):
            for s in sorted(self.spacing_candidates):
                if s > max_spacing + 1e-9:
                    continue
                cand_as = bar_area_mm2(phi) * 1000.0 / s
                if cand_as + 1e-9 < as_req:
                    continue
                cand = RebarSolution(phi, s, cand_as, as_base=0.0, as_add=cand_as, add_text=f"Ø{int(phi)}//{int(s)}", optimization_note=f"optimização wk directa para {label}", n_iterations=2)
                candidates.append(cand)

        # soluções base + reforços coerentes com a malha base
        if self.optimize_rebar and "substituir" not in normalize_text(self.rebar_strategy):
            base = self.base_solution_for_direction(is_vertical_direction)
            ok_base, _ = self.base_solution_is_valid(base, max_spacing)
            if ok_base:
                # base isolada
                if base.as_prov >= as_req:
                    candidates.append(RebarSolution(base.phi, base.spacing, base.as_prov, text_override=base.text, base_text=base.text, add_text="-", as_base=base.as_prov, as_add=0.0, optimization_note=f"base suficiente após verificação wk para {label}", n_iterations=1))
                # reforços intercalados, mesmo espaçamento da base
                for n_layers in (1, 2, 3, 4):
                    for phi in sorted(self.phi_candidates):
                        add_as = n_layers * bar_area_mm2(phi) * 1000.0 / base.spacing
                        final_as = base.as_prov + add_as
                        if final_as + 1e-9 < as_req:
                            continue
                        add_txt = f"{n_layers}Ø{int(phi)}//{int(base.spacing)}" if n_layers > 1 else f"Ø{int(phi)}//{int(base.spacing)}"
                        txt = f"{base.text} + {add_txt}"
                        candidates.append(RebarSolution(min(base.phi, phi), base.spacing, final_as, text_override=txt, base_text=base.text, add_text=add_txt, as_base=base.as_prov, as_add=add_as, optimization_note=f"reforço intercalado por wk para {label}", n_iterations=2 + n_layers))

        valid = []
        fallback = []
        for cand in candidates:
            det = self.crack_check_detail(m_qp, cand, "", "")
            if cand.as_prov + 1e-9 < as_req:
                continue
            key = (cand.as_prov, det["s_eff_mm"], det["phi_control_mm"], cand.text)
            if det["wk_mm"] <= self.wmax_mm + 1e-9 and det["phi_control_mm"] <= det["phi_lim_EC2_mm"] + 1e-9 and det["s_eff_mm"] <= det["s_lim_EC2_mm"] + 1e-9:
                valid.append((key, cand))
            elif det["wk_mm"] <= self.wmax_mm + 1e-9:
                fallback.append((key, cand))
        if valid:
            best = sorted(valid, key=lambda x: x[0])[0][1]
            best.optimization_note = (best.optimization_note + "; " if best.optimization_note else "") + "wk verificado"
            return best
        if fallback:
            best = sorted(fallback, key=lambda x: x[0])[0][1]
            best.optimization_note = (best.optimization_note + "; " if best.optimization_note else "") + "wk cumpre; confirmar limites Ø/s"
            return best
        return sol

    def is_qp_case(self, case_value: object) -> bool:
        if not self.qp_case:
            return False
        return canonical_case_id(case_value) == canonical_case_id(self.qp_case)

    def crack_full_status(self, wk_values: List[float], case_value: object) -> Tuple[str, str]:
        if not self.wk_check_enabled:
            return "Não avaliado", "wk não seleccionado; foi aplicado apenas o controlo simplificado por diâmetro/espaçamento"
        if not self.qp_case:
            return "Verificar", "verificação wk seleccionada, mas o número da combinação quase-permanente não foi indicado"
        if not self.is_qp_case(case_value):
            return "Não avaliado", f"wk só é avaliado para a combinação quase-permanente indicada: {self.qp_case}"
        wk_max = max(wk_values) if wk_values else 0.0
        if wk_max <= self.wmax_mm + 1e-9:
            return "OK", f"combinação quase-permanente {self.qp_case}: wk,max={wk_max:.3f} mm ≤ wmax={self.wmax_mm:.3f} mm"
        return "Verificar", f"combinação quase-permanente {self.qp_case}: wk,max={wk_max:.3f} mm > wmax={self.wmax_mm:.3f} mm"

    def design_one(self, row: pd.Series) -> Dict[str, object]:
        mxx_raw = safe_float(row.get("mxx", 0.0), 0.0)
        myy_raw = safe_float(row.get("myy", 0.0), 0.0)
        mxy_raw = safe_float(row.get("mxy", 0.0), 0.0)
        qxx_raw = safe_float(row.get("qxx", 0.0), 0.0)
        qyy_raw = safe_float(row.get("qyy", 0.0), 0.0)
        mxx = mxx_raw * self.moment_factor
        myy = myy_raw * self.moment_factor
        mxy = mxy_raw * self.moment_factor
        qxx = qxx_raw * self.shear_factor
        qyy = qyy_raw * self.shear_factor
        nxx = safe_float(row.get("nxx", 0.0), 0.0)
        nyy = safe_float(row.get("nyy", 0.0), 0.0)
        nxy = safe_float(row.get("nxy", 0.0), 0.0)

        if self.swap_local_axes:
            mxx, myy = myy, mxx
            qxx, qyy = qyy, qxx
            nxx, nyy = nyy, nxx

        mx_pos, mx_neg, my_pos, my_neg, wa_method_used = self.wood_armer_moments(mxx, myy, mxy)
        m_ref = max(abs(mxx), abs(myy), 1e-9)
        mxy_ratio = abs(mxy) / m_ref

        asx_pos_flex = self.flexural_as_required(mx_pos)
        asx_neg_flex = self.flexural_as_required(mx_neg)
        asy_pos_flex = self.flexural_as_required(my_pos)
        asy_neg_flex = self.flexural_as_required(my_neg)

        asv_min_total, ash_min_total, asv_max_total = self.min_wall_reinf_total()
        # Distribuição dos mínimos de parede por duas faces.
        asv_min_face = asv_min_total / 2.0
        ash_min_face = ash_min_total / 2.0
        # Como os momentos são de placa, reforça-se o mínimo com o mínimo de flexão de lajes/faixa de 1 m.
        as_slab_min_face = self.slab_flexural_min_per_face(12.0)
        as_crack_min_face = self.crack_min_as_per_face()

        if self.local_y_is_vertical:
            asx_min_face = max(ash_min_face, as_slab_min_face, as_crack_min_face)
            asy_min_face = max(asv_min_face, as_slab_min_face, as_crack_min_face)
            sx_max = self.max_spacing_for_direction(False)
            sy_max = self.max_spacing_for_direction(True)
        else:
            asx_min_face = max(asv_min_face, as_slab_min_face, as_crack_min_face)
            asy_min_face = max(ash_min_face, as_slab_min_face, as_crack_min_face)
            sx_max = self.max_spacing_for_direction(True)
            sy_max = self.max_spacing_for_direction(False)

        asx_pos_req = max(asx_pos_flex, asx_min_face)
        asx_neg_req = max(asx_neg_flex, asx_min_face)
        asy_pos_req = max(asy_pos_flex, asy_min_face)
        asy_neg_req = max(asy_neg_flex, asy_min_face)

        x_is_vertical = not self.local_y_is_vertical
        y_is_vertical = self.local_y_is_vertical
        sol_x_pos = self.optimize_rebar_solution(asx_pos_req, sx_max, x_is_vertical, "X+")
        sol_x_neg = self.optimize_rebar_solution(asx_neg_req, sx_max, x_is_vertical, "X-")
        sol_y_pos = self.optimize_rebar_solution(asy_pos_req, sy_max, y_is_vertical, "Y+")
        sol_y_neg = self.optimize_rebar_solution(asy_neg_req, sy_max, y_is_vertical, "Y-")

        # Iteração adicional para esforço transverso: aumenta a armadura na direcção correspondente
        # até VEd <= VRd,c, sempre que tal for possível sem exceder VRd,max.
        d_shear_tmp = self.effective_depth(12.0)
        vrdmax_tmp = self.shear_vrdmax_kN_per_m(d_shear_tmp)
        def shear_target_as(q_abs):
            if q_abs <= 1e-9:
                return 0.0
            best = 0.0
            for a in [i * 50.0 for i in range(1, 801)]:  # até 40000 mm2/m, filtrado depois por As,max
                if self.shear_vrdc_kN_per_m(a, d_shear_tmp) + 1e-9 >= q_abs:
                    return a
                best = a
            return best
        if abs(qxx) > self.shear_vrdc_kN_per_m(max(sol_x_pos.as_prov, sol_x_neg.as_prov), d_shear_tmp) and abs(qxx) <= vrdmax_tmp + 1e-9:
            target = shear_target_as(abs(qxx))
            if target > max(sol_x_pos.as_prov, sol_x_neg.as_prov):
                if sol_x_pos.as_prov <= sol_x_neg.as_prov:
                    sol_x_pos = self.optimize_rebar_solution(max(asx_pos_req, target), sx_max, x_is_vertical, "X+ corte")
                else:
                    sol_x_neg = self.optimize_rebar_solution(max(asx_neg_req, target), sx_max, x_is_vertical, "X- corte")
                sol_x_pos.optimization_note += "; iteração ao corte" if sol_x_pos.as_prov >= target else ""
                sol_x_neg.optimization_note += "; iteração ao corte" if sol_x_neg.as_prov >= target else ""
        if abs(qyy) > self.shear_vrdc_kN_per_m(max(sol_y_pos.as_prov, sol_y_neg.as_prov), d_shear_tmp) and abs(qyy) <= vrdmax_tmp + 1e-9:
            target = shear_target_as(abs(qyy))
            if target > max(sol_y_pos.as_prov, sol_y_neg.as_prov):
                if sol_y_pos.as_prov <= sol_y_neg.as_prov:
                    sol_y_pos = self.optimize_rebar_solution(max(asy_pos_req, target), sy_max, y_is_vertical, "Y+ corte")
                else:
                    sol_y_neg = self.optimize_rebar_solution(max(asy_neg_req, target), sy_max, y_is_vertical, "Y- corte")
                sol_y_pos.optimization_note += "; iteração ao corte" if sol_y_pos.as_prov >= target else ""
                sol_y_neg.optimization_note += "; iteração ao corte" if sol_y_neg.as_prov >= target else ""

        # Iteração específica para fendilhação com a combinação quase-permanente.
        # A optimização usa a armadura final real da face traccionada e, no modo base + reforços,
        # privilegia reforços intercalados com o mesmo espaçamento da armadura base.
        if self.wk_check_enabled and self.is_qp_case(row.get("case", "")):
            sol_x_pos = self.optimize_solution_for_cracking(mx_pos, sol_x_pos, asx_pos_req, sx_max, x_is_vertical, "X+")
            sol_x_neg = self.optimize_solution_for_cracking(mx_neg, sol_x_neg, asx_neg_req, sx_max, x_is_vertical, "X-")
            sol_y_pos = self.optimize_solution_for_cracking(my_pos, sol_y_pos, asy_pos_req, sy_max, y_is_vertical, "Y+")
            sol_y_neg = self.optimize_solution_for_cracking(my_neg, sol_y_neg, asy_neg_req, sy_max, y_is_vertical, "Y-")

        mrd_x_pos = self.mrd_kNm_per_m(sol_x_pos.as_prov, sol_x_pos.phi)
        mrd_x_neg = self.mrd_kNm_per_m(sol_x_neg.as_prov, sol_x_neg.phi)
        mrd_y_pos = self.mrd_kNm_per_m(sol_y_pos.as_prov, sol_y_pos.phi)
        mrd_y_neg = self.mrd_kNm_per_m(sol_y_neg.as_prov, sol_y_neg.phi)

        util_m = max(
            mx_pos / mrd_x_pos if mrd_x_pos > 0 else 0.0,
            mx_neg / mrd_x_neg if mrd_x_neg > 0 else 0.0,
            my_pos / mrd_y_pos if mrd_y_pos > 0 else 0.0,
            my_neg / mrd_y_neg if mrd_y_neg > 0 else 0.0,
        )

        d_shear = self.effective_depth(12.0)
        asx_for_shear = max(sol_x_pos.as_prov, sol_x_neg.as_prov)
        asy_for_shear = max(sol_y_pos.as_prov, sol_y_neg.as_prov)
        vrdc_x = self.shear_vrdc_kN_per_m(asx_for_shear, d_shear)
        vrdc_y = self.shear_vrdc_kN_per_m(asy_for_shear, d_shear)
        vrdmax_x = self.shear_vrdmax_kN_per_m(d_shear)
        vrdmax_y = self.shear_vrdmax_kN_per_m(d_shear)
        util_qx = abs(qxx) / vrdc_x if vrdc_x > 0 else float("inf")
        util_qy = abs(qyy) / vrdc_y if vrdc_y > 0 else float("inf")
        util_qx_max = abs(qxx) / vrdmax_x if vrdmax_x > 0 else float("inf")
        util_qy_max = abs(qyy) / vrdmax_y if vrdmax_y > 0 else float("inf")
        util = max(util_m, util_qx, util_qy)

        # Controlo de As,max para a armadura vertical total, quando o eixo vertical está definido.
        if self.local_y_is_vertical:
            as_vertical_total = sol_y_pos.as_prov + sol_y_neg.as_prov
            max_vert_phi = max(sol_y_pos.phi, sol_y_neg.phi)
        else:
            as_vertical_total = sol_x_pos.as_prov + sol_x_neg.as_prov
            max_vert_phi = max(sol_x_pos.phi, sol_x_neg.phi)
        max_ok = as_vertical_total <= asv_max_total + 1e-9

        # Armadura transversal de ligação entre faces: EC2 9.6.4.
        transverse_links_required = as_vertical_total > 0.02 * self.ac_per_m + 1e-9
        mesh_exception = (max_vert_phi <= 16.0 and self.cover > 2.0 * max_vert_phi)
        transverse_links_note = ""
        crack_items = [
            self.crack_simplified_status(sol_x_pos),
            self.crack_simplified_status(sol_x_neg),
            self.crack_simplified_status(sol_y_pos),
            self.crack_simplified_status(sol_y_neg),
        ]
        crack_x_pos = self.crack_check_detail(mx_pos, sol_x_pos, "+", "X")
        crack_x_neg = self.crack_check_detail(mx_neg, sol_x_neg, "-", "X")
        crack_y_pos = self.crack_check_detail(my_pos, sol_y_pos, "+", "Y")
        crack_y_neg = self.crack_check_detail(my_neg, sol_y_neg, "-", "Y")
        wk_x_pos = float(crack_x_pos["wk_mm"]); sig_x_pos = float(crack_x_pos["sigma_s_MPa"]); sr_x_pos = float(crack_x_pos["sr_max_mm"])
        wk_x_neg = float(crack_x_neg["wk_mm"]); sig_x_neg = float(crack_x_neg["sigma_s_MPa"]); sr_x_neg = float(crack_x_neg["sr_max_mm"])
        wk_y_pos = float(crack_y_pos["wk_mm"]); sig_y_pos = float(crack_y_pos["sigma_s_MPa"]); sr_y_pos = float(crack_y_pos["sr_max_mm"])
        wk_y_neg = float(crack_y_neg["wk_mm"]); sig_y_neg = float(crack_y_neg["sigma_s_MPa"]); sr_y_neg = float(crack_y_neg["sr_max_mm"])
        crack_full_status, crack_full_note = self.crack_full_status([wk_x_pos, wk_x_neg, wk_y_pos, wk_y_neg], row.get("case", ""))
        crack_rank = {"OK": 0, "OK*": 1, "Não avaliado": 1, "Verificar": 2}
        crack_status_simpl = max((st for st, _ in crack_items), key=lambda st: crack_rank.get(st, 2))
        crack_status = max([crack_status_simpl, crack_full_status], key=lambda st: crack_rank.get(st, 2))
        crack_note = " | ".join(sorted(set(note for _, note in crack_items))) + " | " + crack_full_note

        if transverse_links_required:
            transverse_links_note = "requer estribos/ganchos conforme pilares; verificar 9.5.3"
        elif not mesh_exception:
            transverse_links_note = "prever pelo menos 4 estribos/ganchos por m² se a armadura principal estiver junto às faces"
        else:
            transverse_links_note = "pode dispensar armadura transversal se for rede electrossoldada, Ø≤16 e c>2Ø"

        # Verificação preliminar de tensão normal média, se forem fornecidos esforços de membrana.
        # Convenção: NXX/NYY em kN/m, compressão ou tração por metro. 1 kN/m = 1 N/mm.
        n_vertical = nyy if self.local_y_is_vertical else nxx
        sigma_n_mpa = abs(n_vertical) / self.t if self.t > 0 else 0.0
        axial_prelim_util = sigma_n_mpa / self.cp["fcd"] if self.cp["fcd"] > 0 else 0.0

        notes = []
        zero_efforts = max(abs(mxx), abs(myy), abs(mxy)) < 1e-9 and max(abs(qxx), abs(qyy)) < 1e-9
        if self.moment_factor != 1.0 or self.shear_factor != 1.0:
            notes.append(f"unidades convertidas para kNm/m e kN/m: M x{self.moment_factor:g}, Q x{self.shear_factor:g}")
        if self.wk_check_enabled:
            if not self.qp_case:
                notes.append("wk seleccionado, mas falta indicar a combinação quase-permanente")
            elif self.is_qp_case(row.get("case", "")):
                notes.append(f"wk avaliado para a combinação quase-permanente {self.qp_case}")
        else:
            notes.append("wk não seleccionado: fendilhação controlada apenas por diâmetro/espaçamento")
        if axial_prelim_util > 0.60:
            notes.append("tensão normal média elevada; requer verificação de flexão composta/estabilidade")
        elif abs(n_vertical) > 1e-9:
            notes.append("N fornecido: realizada apenas verificação preliminar de tensão média")
        if not max_ok:
            notes.append("As vertical > 0,04 Ac")
        if util_qx > 1.0 or util_qy > 1.0:
            notes.append("VEd > VRd,c; elemento requer verificação adicional ao esforço transverso")
        if util_qx_max > 1.0 or util_qy_max > 1.0:
            notes.append("VEd > VRd,max simplificado; secção não conforme ao limite de esmagamento")
        if transverse_links_required:
            notes.append("As vertical > 0,02 Ac: requer armadura transversal de ligação")
        if mxy_ratio > 0.40:
            notes.append("MXY dominante; confirmar eixos locais e método Wood-Armer")
        if crack_status == "Verificar":
            notes.append("controlo de fendilhação simplificado não satisfeito")
        if self.swap_local_axes:
            notes.append("eixos locais X/Y trocados antes do cálculo")
        if self.optimize_rebar:
            notes.append("optimização base + reforços activada")
        if zero_efforts:
            notes.append("linha sem esforços relevantes")

        if zero_efforts:
            status = "Dados insuficientes"
        elif not max_ok or util_m > 1.0 or util_qx_max > 1.0 or util_qy_max > 1.0:
            status = "Não conforme"
        elif axial_prelim_util > 0.60 or util_qx > 1.0 or util_qy > 1.0 or crack_status == "Verificar":
            status = "Verificar"
        elif mxy_ratio > 0.40 or crack_status == "OK*":
            status = "OK*"
        else:
            status = "OK"

        # Clarificação explícita do estado atribuído.
        reason_items = []
        action_items = []
        if zero_efforts:
            reason_items.append("sem esforços relevantes na linha importada")
            action_items.append("confirmar se a linha deve ser considerada no dimensionamento")
        if not max_ok:
            reason_items.append("As vertical total excede o limite adoptado de 0,04 Ac")
            action_items.append("rever espessura, diâmetros, espaçamentos e eventual necessidade de solução localizada")
        if util_m > 1.0:
            reason_items.append(f"utilização à flexão superior a 1,00 (util_M={util_m:.2f})")
            action_items.append("aumentar a armadura adoptada, reduzir o espaçamento ou rever os esforços de cálculo")
        if util_qx_max > 1.0 or util_qy_max > 1.0:
            reason_items.append("VEd excede VRd,max simplificado")
            action_items.append("aumentar espessura/rever modelo; não adoptar sem verificação específica ao esmagamento")
        elif util_qx > 1.0 or util_qy > 1.0:
            dirs = []
            if util_qx > 1.0:
                dirs.append(f"X (util_Qx={util_qx:.2f})")
            if util_qy > 1.0:
                dirs.append(f"Y (util_Qy={util_qy:.2f})")
            reason_items.append("VEd > VRd,c na direcção " + ", ".join(dirs))
            action_items.append("verificar esforço transverso com maior detalhe, rever espessura ou prever solução específica")
        if crack_status == "Verificar":
            reason_items.append(crack_full_note if self.wk_check_enabled else "controlo simplificado de fendilhação não satisfeito")
            action_items.append("usar combinação quase-permanente de ELS, reduzir diâmetro/espaçamento ou aumentar a armadura")
        elif crack_status == "OK*":
            reason_items.append("controlo de fendilhação preliminar no limite")
            action_items.append("confirmar por wk em ELS se a fendilhação for condicionante")
        if axial_prelim_util > 0.60:
            reason_items.append("tensão normal média elevada detectada")
            action_items.append("realizar verificação N-M/estabilidade fora do âmbito desta versão")
        if transverse_links_required:
            reason_items.append("As vertical > 0,02 Ac")
            action_items.append("prever armadura transversal de ligação entre faces conforme pormenorização aplicável")
        if mxy_ratio > 0.40:
            reason_items.append(f"MXY dominante (|MXY|/max(|MXX|,|MYY|)={mxy_ratio:.2f})")
            action_items.append("confirmar eixos locais do painel e método de transformação de momentos")
        if self.moment_factor != 1.0 or self.shear_factor != 1.0:
            reason_items.append("unidades importadas convertidas antes do cálculo")
            action_items.append("confirmar unidades da tabela de origem")

        if not reason_items:
            if status == "OK":
                reason_items.append("cumpre as verificações implementadas")
                action_items.append("validar hipóteses gerais, eixos locais e dados de entrada")
            elif status == "OK*":
                reason_items.append("cumpre, mas com aviso menor")
                action_items.append("rever notas antes de adoptar em projecto")
            else:
                reason_items.append("estado atribuído por combinação de avisos")
                action_items.append("rever notas de cálculo")

        verification_reason = " | ".join(dict.fromkeys(str(x) for x in reason_items if str(x).strip()))
        recommended_action = " | ".join(dict.fromkeys(str(x) for x in action_items if str(x).strip()))

        return {
            "Panel": row.get("panel", ""),
            "Node": row.get("node", ""),
            "Case": row.get("case", ""),
            "MXX_raw": mxx_raw,
            "MYY_raw": myy_raw,
            "MXY_raw": mxy_raw,
            "QXX_raw": qxx_raw,
            "QYY_raw": qyy_raw,
            "Moment_unit": self.moment_unit,
            "Shear_unit": self.shear_unit,
            "Combo_type": self.combo_type,
            "QP_crack_case": self.qp_case,
            "wk_check_enabled": "Sim" if self.wk_check_enabled else "Não",
            "is_QP_crack_case": "Sim" if self.is_qp_case(row.get("case", "")) else "Não",
            "MXX_kNm_m": mxx,
            "MYY_kNm_m": myy,
            "MXY_kNm_m": mxy,
            "QXX_kN_m": qxx,
            "QYY_kN_m": qyy,
            "NXX_kN_m": nxx,
            "NYY_kN_m": nyy,
            "NXY_kN_m": nxy,
            "Mx+_WA": mx_pos,
            "Mx-_WA": mx_neg,
            "My+_WA": my_pos,
            "My-_WA": my_neg,
            "WA_method": wa_method_used,
            "MXY_ratio": mxy_ratio,
            "Asx+_req_mm2_m": asx_pos_req,
            "Asx-_req_mm2_m": asx_neg_req,
            "Asy+_req_mm2_m": asy_pos_req,
            "Asy-_req_mm2_m": asy_neg_req,
            "Asv_min_total_mm2_m": asv_min_total,
            "Ash_min_total_mm2_m": ash_min_total,
            "As_slab_min_face_mm2_m": as_slab_min_face,
            "As_crack_min_face_mm2_m": as_crack_min_face,
            "sX_max_mm": sx_max,
            "sY_max_mm": sy_max,
            "X+_rebar": sol_x_pos.text,
            "X-_rebar": sol_x_neg.text,
            "Y+_rebar": sol_y_pos.text,
            "Y-_rebar": sol_y_neg.text,
            "X+_base": sol_x_pos.base_text or "-",
            "X-_base": sol_x_neg.base_text or "-",
            "Y+_base": sol_y_pos.base_text or "-",
            "Y-_base": sol_y_neg.base_text or "-",
            "X+_additional": sol_x_pos.add_text or "-",
            "X-_additional": sol_x_neg.add_text or "-",
            "Y+_additional": sol_y_pos.add_text or "-",
            "Y-_additional": sol_y_neg.add_text or "-",
            "X+_optimization": sol_x_pos.optimization_note,
            "X-_optimization": sol_x_neg.optimization_note,
            "Y+_optimization": sol_y_pos.optimization_note,
            "Y-_optimization": sol_y_neg.optimization_note,
            "Optimization_iterations": max(sol_x_pos.n_iterations, sol_x_neg.n_iterations, sol_y_pos.n_iterations, sol_y_neg.n_iterations),
            "Optimization_mode": "Base + reforços" if self.optimize_rebar else "Automática directa",
            "Asx+_prov_mm2_m": sol_x_pos.as_prov,
            "Asx-_prov_mm2_m": sol_x_neg.as_prov,
            "Asy+_prov_mm2_m": sol_y_pos.as_prov,
            "Asy-_prov_mm2_m": sol_y_neg.as_prov,
            "MRd_x+_kNm_m": mrd_x_pos,
            "MRd_x-_kNm_m": mrd_x_neg,
            "MRd_y+_kNm_m": mrd_y_pos,
            "MRd_y-_kNm_m": mrd_y_neg,
            "VRdc_x_kN_m": vrdc_x,
            "VRdc_y_kN_m": vrdc_y,
            "VRdmax_x_kN_m": vrdmax_x,
            "VRdmax_y_kN_m": vrdmax_y,
            "util_M": util_m,
            "util_Qx": util_qx,
            "util_Qy": util_qy,
            "util_Qx_VRdmax": util_qx_max,
            "util_Qy_VRdmax": util_qy_max,
            "util_max": util,
            "As_vertical_total_mm2_m": as_vertical_total,
            "As_vertical_max_mm2_m": asv_max_total,
            "sigma_N_vertical_MPa": sigma_n_mpa,
            "util_N_prelim": axial_prelim_util,
            "Transverse_links_required": "Sim" if transverse_links_required else "Não",
            "Transverse_links_note": transverse_links_note,
            "wk_x+_mm": wk_x_pos,
            "wk_x-_mm": wk_x_neg,
            "wk_y+_mm": wk_y_pos,
            "wk_y-_mm": wk_y_neg,
            "wk_max_mm": max(wk_x_pos, wk_x_neg, wk_y_pos, wk_y_neg),
            "sigma_s_x+_MPa": sig_x_pos,
            "sigma_s_x-_MPa": sig_x_neg,
            "sigma_s_y+_MPa": sig_y_pos,
            "sigma_s_y-_MPa": sig_y_neg,
            "phi_ctrl_x+_mm": crack_x_pos["phi_control_mm"],
            "phi_ctrl_x-_mm": crack_x_neg["phi_control_mm"],
            "phi_ctrl_y+_mm": crack_y_pos["phi_control_mm"],
            "phi_ctrl_y-_mm": crack_y_neg["phi_control_mm"],
            "s_eff_x+_mm": crack_x_pos["s_eff_mm"],
            "s_eff_x-_mm": crack_x_neg["s_eff_mm"],
            "s_eff_y+_mm": crack_y_pos["s_eff_mm"],
            "s_eff_y-_mm": crack_y_neg["s_eff_mm"],
            "phi_lim_x+_EC2_mm": crack_x_pos["phi_lim_EC2_mm"],
            "phi_lim_x-_EC2_mm": crack_x_neg["phi_lim_EC2_mm"],
            "phi_lim_y+_EC2_mm": crack_y_pos["phi_lim_EC2_mm"],
            "phi_lim_y-_EC2_mm": crack_y_neg["phi_lim_EC2_mm"],
            "s_lim_x+_EC2_mm": crack_x_pos["s_lim_EC2_mm"],
            "s_lim_x-_EC2_mm": crack_x_neg["s_lim_EC2_mm"],
            "s_lim_y+_EC2_mm": crack_y_pos["s_lim_EC2_mm"],
            "s_lim_y-_EC2_mm": crack_y_neg["s_lim_EC2_mm"],
            "sr_max_x+_mm": sr_x_pos,
            "sr_max_x-_mm": sr_x_neg,
            "sr_max_y+_mm": sr_y_pos,
            "sr_max_y-_mm": sr_y_neg,
            "Crack_status_x+": crack_x_pos["Status"],
            "Crack_status_x-": crack_x_neg["Status"],
            "Crack_status_y+": crack_y_pos["Status"],
            "Crack_status_y-": crack_y_neg["Status"],
            "Crack_reason_x+": crack_x_pos["Reason"],
            "Crack_reason_x-": crack_x_neg["Reason"],
            "Crack_reason_y+": crack_y_pos["Reason"],
            "Crack_reason_y-": crack_y_neg["Reason"],
            "wmax_mm": self.wmax_mm,
            "Crack_control_status": crack_status,
            "Crack_control_note": crack_note,
            "Status": status,
            "Verification_reason": verification_reason,
            "Recommended_action": recommended_action,
            "Notes": "; ".join(notes),
        }

    def design_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        return pd.DataFrame([self.design_one(r) for _, r in df.iterrows()])


# ============================================================
# GUI
# ============================================================
class WallsEC2App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1220x720")
        self.minsize(900, 560)

        self.df_raw = pd.DataFrame()
        self.df_clean = pd.DataFrame()
        self.df_results = pd.DataFrame()
        self.df_summary = pd.DataFrame()
        self.df_governing = pd.DataFrame()
        self.df_zones = pd.DataFrame()
        self.df_unit_check = pd.DataFrame()
        self.df_detailed = pd.DataFrame()
        self.df_notes = pd.DataFrame()
        self.df_optimization = pd.DataFrame()
        self.df_diagnostic = pd.DataFrame()
        self.df_data_validation = pd.DataFrame()
        self.df_sketches = pd.DataFrame()

        self.var_thickness = tk.StringVar(value="200")
        self.var_cover = tk.StringVar(value="35")
        self.var_concrete = tk.StringVar(value="C30/37")
        self.var_steel = tk.StringVar(value="500")
        self.var_y_vertical = tk.BooleanVar(value=True)
        self.var_swap_axes = tk.BooleanVar(value=False)
        self.var_crack_spacing = tk.BooleanVar(value=True)
        self.var_crack_check = tk.BooleanVar(value=False)
        self.var_qp_case = tk.StringVar(value="")
        self.var_combo_type = tk.StringVar(value="ELU")
        self.var_wmax = tk.StringVar(value="0.30")
        self.var_moment_unit = tk.StringVar(value="kNm/m")
        self.var_shear_unit = tk.StringVar(value="kN/m")
        self.var_phi_min = tk.StringVar(value="8")
        self.var_phi_max = tk.StringVar(value="16")
        self.var_optimize_rebar = tk.BooleanVar(value=True)
        self.var_rebar_strategy = tk.StringVar(value="Base + reforços")
        self.var_base_v_phi = tk.StringVar(value="10")
        self.var_base_v_spacing = tk.StringVar(value="200")
        self.var_base_h_phi = tk.StringVar(value="8")
        self.var_base_h_spacing = tk.StringVar(value="200")
        self.var_wa_method = tk.StringVar(value="Conservativo |MXY|")
        self.var_reduce = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="Cole a tabela de esforços ou importe um ficheiro.")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_text_var = tk.StringVar(value="0%")
        self.progress_var.trace_add("write", lambda *args: self.progress_text_var.set(f"{self.progress_var.get():.0f}%"))

        self._build_ui()

    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        # ---- aparência geral ----
        bg = style.lookup("TFrame", "background") or "#f3f5f7"
        self.configure(background=bg)
        style.configure("TLabelframe", padding=8)
        style.configure("TLabelframe.Label", font=("Segoe UI", 9, "bold"))
        style.configure("TButton", padding=(8, 6))
        style.configure("Primary.TButton", padding=(10, 8))
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("TNotebook.Tab", padding=(10, 6), font=("Segoe UI", 9))
        style.configure("Header.TLabel", font=("Segoe UI Semibold", 11))
        style.configure("Subtle.TLabel", font=("Segoe UI", 8), foreground="#5f6b7a")
        style.configure("Status.TLabel", font=("Segoe UI", 9))

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)

        root = ttk.Frame(self, padding=8)
        root.grid(row=0, column=0, sticky="nsew")
        root.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)

        paned = ttk.Panedwindow(root, orient="horizontal")
        paned.grid(row=0, column=0, sticky="nsew")

        # ---- barra lateral scrollable, mais adequada a ecrãs menores ----
        sidebar_host = ttk.Frame(paned, width=355)
        sidebar_host.pack_propagate(False)
        sidebar_host.rowconfigure(0, weight=1)
        sidebar_host.columnconfigure(0, weight=1)

        self.sidebar_canvas = tk.Canvas(sidebar_host, highlightthickness=0, borderwidth=0, background=bg)
        sb_y = ttk.Scrollbar(sidebar_host, orient="vertical", command=self.sidebar_canvas.yview)
        self.sidebar_canvas.configure(yscrollcommand=sb_y.set)
        self.sidebar_canvas.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")

        sidebar_inner = ttk.Frame(self.sidebar_canvas, padding=(0, 0, 6, 0))
        sidebar_window = self.sidebar_canvas.create_window((0, 0), window=sidebar_inner, anchor="nw")

        def _sync_sidebar(_event=None):
            self.sidebar_canvas.configure(scrollregion=self.sidebar_canvas.bbox("all"))
            width = self.sidebar_canvas.winfo_width()
            self.sidebar_canvas.itemconfigure(sidebar_window, width=width)

        sidebar_inner.bind("<Configure>", _sync_sidebar)
        self.sidebar_canvas.bind("<Configure>", _sync_sidebar)

        def _on_mousewheel(event):
            try:
                self.sidebar_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        self.sidebar_canvas.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        self.sidebar_canvas.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

        right = ttk.Frame(paned)
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        paned.add(sidebar_host, weight=0)
        paned.add(right, weight=1)

        # ---- rodapé sempre visível ----
        bottom = ttk.Frame(self, padding=(8, 4, 8, 8))
        bottom.grid(row=1, column=0, sticky="ew")
        bottom.columnconfigure(1, weight=1)
        ttk.Label(bottom, text="Estado:", style="Header.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(bottom, textvariable=self.status_var, style="Status.TLabel").grid(row=0, column=1, sticky="ew")
        self.progress_bottom = ttk.Progressbar(bottom, variable=self.progress_var, maximum=100, length=260)
        self.progress_bottom.grid(row=0, column=2, sticky="e", padx=(12, 6))
        ttk.Label(bottom, textvariable=self.progress_text_var, width=6, anchor="e").grid(row=0, column=3, sticky="e")

        # ---- cabeçalho compacto ----
        hero = ttk.LabelFrame(sidebar_inner, text="WallsEC2")
        hero.pack(fill="x", pady=(0, 8))
        program_link = ttk.Label(hero, text="WallsEC2", style="Header.TLabel", cursor="hand2")
        program_link.pack(anchor="w")
        program_link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
        ttk.Label(hero, text="Dimensionamento de paredes de betão armado (EC2)", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
        ttk.Label(
            hero,
            text="Ferramenta para o dimensionamento de paredes de B.A. segundo o EC2, com verificação de armaduras, corte, fendilhação e relatórios .xlsx/.pdf",
            style="Subtle.TLabel",
            wraplength=310,
            justify="left",
        ).pack(anchor="w", pady=(2, 0))

        # ---- secção 1: geometria e materiais ----
        basic = ttk.LabelFrame(sidebar_inner, text="1. Geometria e materiais")
        basic.pack(fill="x", pady=(0, 8))
        self._add_label_entry(basic, "Espessura [mm]", self.var_thickness, 0)
        self._add_label_entry(basic, "Recobrimento [mm]", self.var_cover, 1)
        ttk.Label(basic, text="Classe de Betão").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(
            basic, textvariable=self.var_concrete, width=12, state="readonly",
            values=["C20/25", "C25/30", "C30/37", "C35/45", "C40/50", "C45/55", "C50/60"]
        ).grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(basic, text="Classe de Aço [MPa]").grid(row=3, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(basic, textvariable=self.var_steel, width=12, state="readonly", values=["400", "500"]).grid(
            row=3, column=1, sticky="ew", padx=6, pady=4
        )
        basic.columnconfigure(1, weight=1)

        # ---- secção 2: modelação, verificações e unidades ----
        checks = ttk.LabelFrame(sidebar_inner, text="2. Modelação e verificações")
        checks.pack(fill="x", pady=(0, 8))
        ttk.Checkbutton(checks, text="Eixo local Y = vertical", variable=self.var_y_vertical).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=6, pady=3
        )
        ttk.Checkbutton(checks, text="Trocar eixos locais X ↔ Y antes do cálculo", variable=self.var_swap_axes).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=6, pady=3
        )
        ttk.Label(checks, text="Esforços").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(checks, textvariable=self.var_wa_method, state="readonly",
                     values=["Conservativo |MXY|", "Momentos principais"]).grid(
            row=2, column=1, sticky="ew", padx=6, pady=4
        )
        ttk.Checkbutton(
            checks, text="Limitar espaçamento a 250 mm", variable=self.var_crack_spacing
        ).grid(row=3, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        ttk.Checkbutton(
            checks, text="Verificar wk com combinação quase-permanente", variable=self.var_crack_check
        ).grid(row=4, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        ttk.Label(checks, text="Comb. quase-perm.").grid(row=5, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(checks, textvariable=self.var_qp_case, width=14).grid(row=5, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(checks, text="wmax [mm]").grid(row=6, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(checks, textvariable=self.var_wmax, width=14).grid(row=6, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(checks, text="Unid. momentos").grid(row=7, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(checks, textvariable=self.var_moment_unit, state="readonly",
                     values=["kNm/m", "Nmm/mm", "Nm/m", "kNmm/m"]).grid(
            row=7, column=1, sticky="ew", padx=6, pady=4
        )
        ttk.Label(checks, text="Unid. corte").grid(row=8, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(checks, textvariable=self.var_shear_unit, state="readonly", values=["kN/m", "N/mm", "N/m"]).grid(
            row=8, column=1, sticky="ew", padx=6, pady=4
        )
        checks.columnconfigure(1, weight=1)

        # ---- secção 3: armaduras e optimização ----
        rebar = ttk.LabelFrame(sidebar_inner, text="3. Armaduras e optimização")
        rebar.pack(fill="x", pady=(0, 8))
        ttk.Label(rebar, text="Ø min / Ø max [mm]").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        phi_frame = ttk.Frame(rebar)
        phi_frame.grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        ttk.Entry(phi_frame, textvariable=self.var_phi_min, width=5).pack(side="left")
        ttk.Label(phi_frame, text="  /  ").pack(side="left")
        ttk.Entry(phi_frame, textvariable=self.var_phi_max, width=5).pack(side="left")
        ttk.Checkbutton(rebar, text="Optimizar armadura", variable=self.var_optimize_rebar).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=6, pady=3
        )
        ttk.Label(rebar, text="Estratégia").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(rebar, textvariable=self.var_rebar_strategy, state="readonly",
                     values=["Base + reforços", "Substituir base"]).grid(
            row=2, column=1, sticky="ew", padx=6, pady=4
        )
        ttk.Label(rebar, text="Base vertical Ø//s").grid(row=3, column=0, sticky="w", padx=6, pady=4)
        basev = ttk.Frame(rebar)
        basev.grid(row=3, column=1, sticky="ew", padx=6, pady=4)
        ttk.Entry(basev, textvariable=self.var_base_v_phi, width=5).pack(side="left")
        ttk.Label(basev, text=" // ").pack(side="left")
        ttk.Entry(basev, textvariable=self.var_base_v_spacing, width=6).pack(side="left")
        ttk.Label(rebar, text="Base horizontal Ø//s").grid(row=4, column=0, sticky="w", padx=6, pady=4)
        baseh = ttk.Frame(rebar)
        baseh.grid(row=4, column=1, sticky="ew", padx=6, pady=4)
        ttk.Entry(baseh, textvariable=self.var_base_h_phi, width=5).pack(side="left")
        ttk.Label(baseh, text=" // ").pack(side="left")
        ttk.Entry(baseh, textvariable=self.var_base_h_spacing, width=6).pack(side="left")
        ttk.Checkbutton(rebar, text="Reduzir para casos governantes", variable=self.var_reduce).grid(
            row=5, column=0, columnspan=2, sticky="w", padx=6, pady=3
        )
        rebar.columnconfigure(1, weight=1)

        # ---- acções ----
        actions = ttk.LabelFrame(sidebar_inner, text="4. Ações")
        actions.pack(fill="x", pady=(0, 8))
        btns = [
            ("Colar área de transferência", self.paste_clipboard),
            ("Importar .xlsx/.csv", self.import_file),
            ("Calcular", self.run_design),
            ("Exportar .xlsx", self.export_excel),
            ("Relatório .pdf", self.export_pdf_report),
            ("Exportar .csv", self.export_csv),
            ("Guardar configuração", self.save_config),
            ("Carregar configuração", self.load_config),
        ]
        for i, (txt, cmd) in enumerate(btns):
            style_name = "Primary.TButton" if txt == "Calcular" else "TButton"
            ttk.Button(actions, text=txt, command=cmd, style=style_name).grid(
                row=i // 2, column=i % 2, sticky="ew", padx=4, pady=4
            )
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)

        # ---- estado rápido ----
        progress_box = ttk.LabelFrame(sidebar_inner, text="5. Estado")
        progress_box.pack(fill="x", pady=(0, 8))
        ttk.Label(progress_box, textvariable=self.status_var, wraplength=310, justify="left").pack(fill="x", padx=6, pady=(4, 2))
        self.progress = ttk.Progressbar(progress_box, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", padx=6, pady=(2, 2))
        ttk.Label(progress_box, textvariable=self.progress_text_var, anchor="e").pack(fill="x", padx=6, pady=(0, 4))

        # ---- notas rápidas mais compactas ----
        quick = ttk.LabelFrame(sidebar_inner, text="6. Notas rápidas")
        quick.pack(fill="x", pady=(0, 8))
        quick_text = (
            "• Paredes tratadas como painéis/placas.\n"
            "• Entrada principal: MXX, MYY, MXY, QXX, QYY.\n"
            "• Faixas de 1 m; Wood-Armer simplificado.\n"
            "• Pormenorização mínima EC2 9.6.\n"
            "• Sem NXX/NYY/NXY não há verificação global de compressão/2ª ordem."
        )
        ttk.Label(quick, text=quick_text, wraplength=310, justify="left").pack(fill="x", padx=6, pady=6)

        # ---- painel principal com tabs curtas ----
        nb = ttk.Notebook(right)
        nb.grid(row=0, column=0, sticky="nsew")
        try:
            nb.enable_traversal()
        except Exception:
            pass

        self.tab_instructions = ttk.Frame(nb)
        self.tab_paste = ttk.Frame(nb)
        self.tab_input = ttk.Frame(nb)
        self.tab_results = ttk.Frame(nb)
        self.tab_summary = ttk.Frame(nb)
        self.tab_governing = ttk.Frame(nb)
        self.tab_zones = ttk.Frame(nb)
        self.tab_optimization = ttk.Frame(nb)
        self.tab_diagnostic = ttk.Frame(nb)
        self.tab_validation = ttk.Frame(nb)
        self.tab_sketches = ttk.Frame(nb)
        self.tab_unit = ttk.Frame(nb)
        self.tab_detailed = ttk.Frame(nb)
        self.tab_notes = ttk.Frame(nb)
        tab_defs = [
            (self.tab_instructions, "Instruções"),
            (self.tab_paste, "Colar"),
            (self.tab_input, "Tabela"),
            (self.tab_results, "Resultados"),
            (self.tab_summary, "Resumo"),
            (self.tab_governing, "Armaduras"),
            (self.tab_zones, "Zonas"),
            (self.tab_optimization, "Optimização"),
            (self.tab_diagnostic, "Diagnóstico"),
            (self.tab_validation, "Validação"),
            (self.tab_sketches, "Croquis"),
            (self.tab_unit, "Unidades"),
            (self.tab_detailed, "Detalhe"),
            (self.tab_notes, "Notas EC2"),
        ]
        for frame, title in tab_defs:
            nb.add(frame, text=title)

        self._build_instructions_tab(self.tab_instructions)
        self._build_paste_tab(self.tab_paste)
        self.tree_input = self._make_tree(self.tab_input)
        self.tree_results = self._make_tree(self.tab_results)
        self.tree_summary = self._make_tree(self.tab_summary)
        self.tree_governing = self._make_tree(self.tab_governing)
        self.tree_zones = self._make_tree(self.tab_zones)
        self.tree_optimization = self._make_tree(self.tab_optimization)
        self.tree_diagnostic = self._make_tree(self.tab_diagnostic)
        self.tree_validation = self._make_tree(self.tab_validation)
        self.tree_sketches = self._make_tree(self.tab_sketches)
        self.tree_unit = self._make_tree(self.tab_unit)
        self.tree_detailed = self._make_tree(self.tab_detailed)
        self.tree_notes = self._make_tree(self.tab_notes)

        # posição inicial da divisória pensada para ecrãs menores
        self.after(100, lambda: paned.sashpos(0, 370))

    def _add_label_entry(self, parent, label, var, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(parent, textvariable=var, width=14).grid(row=row, column=1, sticky="ew", padx=6, pady=4)

    def _build_instructions_tab(self, parent):
        outer = ttk.Frame(parent, padding=10)
        outer.pack(fill="both", expand=True)
        outer.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)

        ttk.Label(
            outer,
            text="Instruções de importação da tabela de esforços",
            style="Header.TLabel",
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        frame = ttk.Frame(outer)
        frame.grid(row=1, column=0, sticky="nsew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        txt = tk.Text(frame, wrap="word", height=22, undo=False)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        txt.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        instructions = (
            "OBJECTIVO\n"
            "Esta ferramenta dimensiona armadura distribuída de paredes/painéis de betão armado a partir de esforços por metro.\n\n"
            "COMO USAR\n"
            "1. Copiar a tabela de esforços do software de análise ou de uma folha de cálculo.\n"
            "2. Abrir a aba 'Colar'.\n"
            "3. Colar a tabela.\n"
            "4. Clicar em 'Ler tabela colada'.\n"
            "5. Confirmar unidades, eixos locais e combinação quase-permanente, quando aplicável.\n"
            "6. Clicar em 'Calcular'.\n\n"
            "COLUNAS RECOMENDADAS\n"
            "A tabela deve conter, sempre que possível:\n"
            "- Panel ou Painel\n"
            "- Node ou Nó\n"
            "- Case ou Caso/Combinação\n"
            "- MXX\n"
            "- MYY\n"
            "- MXY\n"
            "- QXX\n"
            "- QYY\n\n"
            "UNIDADES ESPERADAS\n"
            "- MXX, MYY e MXY: normalmente em kNm/m.\n"
            "- QXX e QYY: normalmente em kN/m.\n"
            "As unidades podem ser alteradas no painel lateral.\n\n"
            "TABELA TIPO\n"
            "Panel\tNode\tCase\tMXX\tMYY\tMXY\tQXX\tQYY\n"
            "43\t49\t101\t-12.40\t3.80\t1.25\t18.50\t6.20\n"
            "43\t49\t302 (QP)\t-5.10\t1.40\t0.52\t7.20\t2.10\n"
            "44\t51\t101\t8.75\t-2.60\t0.90\t12.40\t5.80\n\n"
            "FORMATO ALTERNATIVO ACEITE\n"
            "Também é aceite uma coluna única do tipo Panel/Node/Case, por exemplo:\n\n"
            "Panel/Node/Case\tMXX\tMYY\tMXY\tQXX\tQYY\n"
            "43/49/101\t-12.40\t3.80\t1.25\t18.50\t6.20\n"
            "43/49/302 (QP)\t-5.10\t1.40\t0.52\t7.20\t2.10\n\n"
            "NOTAS IMPORTANTES\n"
            "- Confirmar a orientação dos eixos locais dos painéis antes de adoptar as armaduras.\n"
            "- A face +/− depende da normal local do painel.\n"
            "- Para verificação explícita de wk, indicar o número da combinação quase-permanente.\n"
            "- Sem esforços normais/membrana, a ferramenta não verifica compressão global, estabilidade ou segunda ordem.\n"
            "- O relatório PDF é resumido; o Excel exportado mantém os resultados completos para auditoria.\n"
        )
        txt.insert("1.0", instructions)
        txt.config(state="disabled")

    def _build_paste_tab(self, parent):
        top = ttk.Frame(parent, padding=6)
        top.pack(fill="x")
        ttk.Label(top, text="Cole aqui a tabela de esforços e clique em 'Ler tabela colada'.").pack(side="left")
        ttk.Button(top, text="Ler tabela colada", command=self.load_from_textbox).pack(side="right")
        ttk.Button(top, text="Limpar", command=lambda: self.txt_paste.delete("1.0", "end")).pack(side="right", padx=(0, 6))

        frame = ttk.Frame(parent, padding=(6, 0, 6, 6))
        frame.pack(fill="both", expand=True)
        self.txt_paste = tk.Text(frame, wrap="none", undo=True)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.txt_paste.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.txt_paste.xview)
        self.txt_paste.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.txt_paste.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    def _make_tree(self, parent):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tree

    def show_df(self, tree: ttk.Treeview, df: pd.DataFrame):
        tree.delete(*tree.get_children())
        if df is None or df.empty:
            tree["columns"] = []
            return
        cols = list(df.columns)
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(90, min(180, len(str(c)) * 9)), anchor="center")
        for _, row in df.head(MAX_PREVIEW_ROWS).iterrows():
            vals = []
            for c in cols:
                v = row[c]
                if isinstance(v, float):
                    vals.append("" if not math.isfinite(v) else f"{v:.3f}")
                else:
                    vals.append(str(v))
            tree.insert("", "end", values=vals)

    def load_from_textbox(self):
        text = self.txt_paste.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Aviso", "Cole primeiro a tabela na caixa de texto.")
            return
        df = parse_pasted_table(text)
        if df.empty:
            messagebox.showwarning("Aviso", "A tabela colada não foi reconhecida.")
            return
        self.load_df(df)

    def paste_clipboard(self):
        try:
            text = self.clipboard_get()
        except Exception:
            messagebox.showwarning("Aviso", "Não foi possível ler a área de transferência.")
            return
        df = parse_pasted_table(text)
        if df.empty:
            messagebox.showwarning("Aviso", "A tabela colada não foi reconhecida.")
            return
        self.txt_paste.delete("1.0", "end")
        self.txt_paste.insert("1.0", text)
        self.load_df(df)

    def import_file(self):
        path = filedialog.askopenfilename(
            title="Importar tabela",
            filetypes=[("Excel", "*.xlsx *.xls"), ("CSV", "*.csv"), ("Todos", "*.*")],
        )
        if not path:
            return
        try:
            if path.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(path, dtype=str)
            else:
                # tenta separadores correntes
                try:
                    df = pd.read_csv(path, dtype=str)
                except Exception:
                    df = pd.read_csv(path, sep=";", dtype=str)
            self.load_df(df)
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def load_df(self, df: pd.DataFrame):
        self.df_raw = df.copy()
        self.df_clean = clean_dataframe(df)
        self.df_results = pd.DataFrame()
        self.df_summary = pd.DataFrame()
        self.df_governing = pd.DataFrame()
        self.df_zones = pd.DataFrame()
        self.df_unit_check = pd.DataFrame()
        self.df_detailed = pd.DataFrame()
        self.df_notes = pd.DataFrame()
        self.df_optimization = pd.DataFrame()
        self.df_diagnostic = pd.DataFrame()
        self.df_data_validation = pd.DataFrame()
        self.df_sketches = pd.DataFrame()
        self.show_df(self.tree_input, self.df_clean)
        self.show_df(self.tree_results, self.df_results)
        self.show_df(self.tree_summary, self.df_summary)
        self.show_df(self.tree_governing, self.df_governing)
        self.show_df(self.tree_zones, self.df_zones)
        self.show_df(self.tree_optimization, self.df_optimization)
        self.show_df(self.tree_diagnostic, self.df_diagnostic)
        self.show_df(self.tree_validation, self.df_data_validation)
        self.show_df(self.tree_sketches, self.df_sketches)
        self.show_df(self.tree_unit, self.df_unit_check)
        self.show_df(self.tree_detailed, self.df_detailed)
        self.show_df(self.tree_notes, self.df_notes)
        self.progress_var.set(0.0)
        self.status_var.set(f"Tabela carregada: {len(self.df_clean)} linhas.")

    def validate_inputs(self) -> Optional[str]:
        if self.df_clean is None or self.df_clean.empty:
            return "Cole ou importe uma tabela de esforços primeiro."
        t = safe_float(self.var_thickness.get(), 0.0)
        c = safe_float(self.var_cover.get(), 0.0)
        if t <= 0:
            return "Espessura inválida."
        if c <= 0 or c >= 0.45 * t:
            return "Recobrimento inválido para a espessura indicada."
        return None

    def build_unit_check(self, df_clean: pd.DataFrame, designer: WallDesigner) -> pd.DataFrame:
        if df_clean is None or df_clean.empty:
            return pd.DataFrame()
        vals_m = []
        vals_q = []
        for c in ["mxx", "myy", "mxy"]:
            if c in df_clean.columns:
                vals_m.extend([abs(v) * designer.moment_factor for v in df_clean[c].dropna().map(lambda x: safe_float(x, 0.0)).tolist()])
        for c in ["qxx", "qyy"]:
            if c in df_clean.columns:
                vals_q.extend([abs(v) * designer.shear_factor for v in df_clean[c].dropna().map(lambda x: safe_float(x, 0.0)).tolist()])
        max_m = max(vals_m) if vals_m else 0.0
        max_q = max(vals_q) if vals_q else 0.0
        notes = []
        if max_m > 1000.0:
            notes.append("Momentos convertidos muito elevados para paredes correntes; confirmar unidades e sinal dos esforços.")
        if 0.0 < max_m < 0.01:
            notes.append("Momentos muito baixos; confirmar se a unidade não deveria ser kNm/m.")
        if max_q > 1500.0:
            notes.append("Esforços transversos muito elevados para paredes correntes; confirmar unidades.")
        if not notes:
            notes.append("Sem anomalias óbvias por ordem de grandeza. Esta verificação não substitui a confirmação das unidades na origem dos resultados.")
        return pd.DataFrame([
            {"Grandeza": "Momentos MXX/MYY/MXY", "Unidade selecionada": designer.moment_unit, "Factor aplicado": designer.moment_factor, "Máximo convertido": max_m, "Unidade interna": "kNm/m", "Nota": notes[0]},
            {"Grandeza": "Corte QXX/QYY", "Unidade selecionada": designer.shear_unit, "Factor aplicado": designer.shear_factor, "Máximo convertido": max_q, "Unidade interna": "kN/m", "Nota": " | ".join(notes)},
            {"Grandeza": "Fendilhação", "Unidade selecionada": "wk seleccionado" if designer.wk_check_enabled else "controlo simplificado", "Factor aplicado": "-", "Máximo convertido": "-", "Unidade interna": "-", "Nota": f"wk só é avaliado para a combinação quase-permanente indicada: {designer.qp_case or 'não indicada'}. Sem selecção, fica o controlo por diâmetro/espaçamento."},
        ])


    def build_data_validation(self, df_clean: pd.DataFrame, designer: WallDesigner) -> pd.DataFrame:
        rows = []
        if df_clean is None or df_clean.empty:
            return pd.DataFrame([{"Categoria": "Tabela", "Item": "linhas", "Estado": "Não conforme", "Resultado": "0", "Nota": "Sem dados importados."}])

        required = ["panel", "case", "mxx", "myy", "mxy", "qxx", "qyy"]
        for c in required:
            estado = "OK" if c in df_clean.columns else "Não conforme"
            rows.append({"Categoria": "Colunas", "Item": c, "Estado": estado, "Resultado": "presente" if estado == "OK" else "em falta", "Nota": "coluna reconhecida" if estado == "OK" else "corrigir cabeçalho ou formato de importação"})

        n_lines = len(df_clean)
        n_panels = df_clean["panel"].astype(str).nunique() if "panel" in df_clean.columns else 0
        n_cases = df_clean["case_id"].astype(str).nunique() if "case_id" in df_clean.columns else 0
        rows.append({"Categoria": "Tabela", "Item": "linhas", "Estado": "OK" if n_lines > 0 else "Não conforme", "Resultado": n_lines, "Nota": "número de linhas importadas"})
        rows.append({"Categoria": "Tabela", "Item": "painéis", "Estado": "OK" if n_panels > 0 else "Verificar", "Resultado": n_panels, "Nota": "painéis distintos detectados"})
        rows.append({"Categoria": "Tabela", "Item": "combinações", "Estado": "OK" if n_cases > 0 else "Verificar", "Resultado": n_cases, "Nota": "casos/combinações distintos detectados"})

        if designer.wk_check_enabled:
            qp = designer.qp_case
            if not qp:
                rows.append({"Categoria": "ELS", "Item": "combinação quase-permanente", "Estado": "Não conforme", "Resultado": "não indicada", "Nota": "wk seleccionado sem número de combinação quase-permanente"})
            else:
                mask = df_clean["case"].map(designer.is_qp_case) if "case" in df_clean.columns else pd.Series(False, index=df_clean.index)
                n_qp = int(mask.sum())
                panels_qp = df_clean.loc[mask, "panel"].astype(str).nunique() if n_qp else 0
                missing_panels = []
                if "panel" in df_clean.columns and n_qp:
                    all_p = set(df_clean["panel"].astype(str))
                    qp_p = set(df_clean.loc[mask, "panel"].astype(str))
                    missing_panels = sorted(all_p - qp_p)
                estado = "OK" if n_qp > 0 and not missing_panels else ("Verificar" if n_qp > 0 else "Não conforme")
                rows.append({"Categoria": "ELS", "Item": "combinação quase-permanente", "Estado": estado, "Resultado": f"{qp}: {n_qp} linhas; {panels_qp}/{n_panels} painéis", "Nota": ("encontrada em todos os painéis" if not missing_panels and n_qp else f"painéis sem QP: {', '.join(missing_panels[:12])}" if missing_panels else "combinação não encontrada")})

        # Coerência básica de ordem de grandeza, após conversão.
        max_m = max([df_clean[c].abs().max() * designer.moment_factor for c in ["mxx", "myy", "mxy"] if c in df_clean.columns] or [0])
        max_q = max([df_clean[c].abs().max() * designer.shear_factor for c in ["qxx", "qyy"] if c in df_clean.columns] or [0])
        rows.append({"Categoria": "Unidades", "Item": "momentos", "Estado": "Verificar" if max_m > 1000 or (0 < max_m < 0.01) else "OK", "Resultado": f"max |M| = {max_m:.3f} kNm/m", "Nota": "confirmar unidades se valor estiver fora da ordem corrente"})
        rows.append({"Categoria": "Unidades", "Item": "corte", "Estado": "Verificar" if max_q > 1500 else "OK", "Resultado": f"max |Q| = {max_q:.3f} kN/m", "Nota": "confirmar unidades se valor estiver fora da ordem corrente"})
        rows.append({"Categoria": "Eixos locais", "Item": "orientação", "Estado": "OK*" if designer.swap_local_axes else "Verificar", "Resultado": "X/Y trocados" if designer.swap_local_axes else ("Y local = vertical" if designer.local_y_is_vertical else "X local = vertical"), "Nota": "confirmar orientação dos eixos locais e da normal dos painéis antes de desenhar armaduras"})
        return pd.DataFrame(rows)

    def build_diagnostic(self, results: pd.DataFrame, validation: pd.DataFrame) -> pd.DataFrame:
        rows = []
        def add(panel, sev, problem, action, source="Cálculo"):
            rows.append({"Painel": panel, "Severidade": sev, "Problema": problem, "Acção recomendada": action, "Origem": source})

        if validation is not None and not validation.empty:
            for _, r in validation.iterrows():
                estado = str(r.get("Estado", ""))
                if estado in ("Verificar", "Não conforme", "OK*"):
                    sev = "Alta" if estado == "Não conforme" else ("Média" if estado == "Verificar" else "Baixa")
                    add("-", sev, f"{r.get('Categoria','')} - {r.get('Item','')}: {r.get('Resultado','')}", r.get("Nota", ""), source="Validação da tabela")

        if results is None or results.empty:
            return pd.DataFrame(rows)
        for panel, g in results.groupby("Panel", dropna=False):
            # Estado mais gravoso por painel
            statuses = set(str(x) for x in g.get("Status", pd.Series(dtype=str)).dropna())
            if "Não conforme" in statuses:
                r = g.loc[g["util_max"].fillna(-1).idxmax()]
                add(panel, "Alta", r.get("Verification_reason", "Não conforme"), r.get("Recommended_action", "rever dimensionamento"))
            elif "Verificar" in statuses:
                r = g.loc[g["util_max"].fillna(-1).idxmax()]
                add(panel, "Média", r.get("Verification_reason", "Verificar"), r.get("Recommended_action", "validar manualmente"))
            elif "OK*" in statuses:
                r = g.loc[g["util_max"].fillna(-1).idxmax()]
                add(panel, "Baixa", r.get("Verification_reason", "OK com aviso"), r.get("Recommended_action", "confirmar notas"))

            if "MXY_ratio" in g.columns and g["MXY_ratio"].max() > 0.40:
                add(panel, "Média", f"MXY dominante: ratio máximo = {g['MXY_ratio'].max():.2f}", "confirmar eixos locais, orientação da normal e método adoptado para MXY")
            if "is_QP_crack_case" in g.columns and "Sim" not in set(g["is_QP_crack_case"].astype(str)):
                # Só é problema se wk foi activado; neste caso a coluna QP existe, mas o caso não apareceu no painel.
                if "wk_check_enabled" in g.columns and "Sim" in set(g["wk_check_enabled"].astype(str)):
                    add(panel, "Média", "sem linha da combinação quase-permanente para verificação wk", "exportar/colar também a combinação quase-permanente de ELS")
        if not rows:
            rows.append({"Painel": "-", "Severidade": "Informação", "Problema": "sem problemas críticos detectados", "Acção recomendada": "validar hipóteses gerais, eixos locais e modelo estrutural", "Origem": "Diagnóstico"})
        order = {"Alta": 0, "Média": 1, "Baixa": 2, "Informação": 3}
        out = pd.DataFrame(rows)
        out["_ord"] = out["Severidade"].map(order).fillna(9)
        return out.sort_values(["_ord", "Painel"]).drop(columns=["_ord"]).reset_index(drop=True)

    def build_rebar_sketches(self, zones: pd.DataFrame) -> pd.DataFrame:
        if zones is None or zones.empty:
            return pd.DataFrame()
        rows = []
        for _, r in zones.iterrows():
            sketch = (
                f"Zona {r.get('Zona','')} | Painéis: {r.get('Painéis','')}\n"
                f"Face +: X = {r.get('X+','-')} | Y = {r.get('Y+','-')}\n"
                f"Face -: X = {r.get('X-','-')} | Y = {r.get('Y-','-')}\n"
                f"Nota: armaduras adicionais devem ser intercaladas com a armadura base, mantendo o mesmo espaçamento da base."
            )
            rows.append({"Zona": r.get("Zona", ""), "Painéis": r.get("Painéis", ""), "Croquis_textual": sketch, "Nota_desenho": "Face +/− depende da normal local do painel; confirmar antes de desenhar."})
        return pd.DataFrame(rows)

    def build_detailed_check(self, results: pd.DataFrame) -> pd.DataFrame:
        if results is None or results.empty:
            return pd.DataFrame()
        cols = [
            "Panel", "Node", "Case", "Combo_type", "MXX_kNm_m", "MYY_kNm_m", "MXY_kNm_m",
            "Mx+_WA", "Mx-_WA", "My+_WA", "My-_WA", "WA_method",
            "Asx+_req_mm2_m", "Asx-_req_mm2_m", "Asy+_req_mm2_m", "Asy-_req_mm2_m",
            "As_slab_min_face_mm2_m", "As_crack_min_face_mm2_m", "X+_rebar", "X-_rebar", "Y+_rebar", "Y-_rebar",
            "MRd_x+_kNm_m", "MRd_x-_kNm_m", "MRd_y+_kNm_m", "MRd_y-_kNm_m",
            "QXX_kN_m", "QYY_kN_m", "VRdc_x_kN_m", "VRdc_y_kN_m", "VRdmax_x_kN_m", "VRdmax_y_kN_m",
            "wk_x+_mm", "wk_x-_mm", "wk_y+_mm", "wk_y-_mm", "wk_max_mm", "wmax_mm",
            "sigma_s_x+_MPa", "sigma_s_x-_MPa", "sigma_s_y+_MPa", "sigma_s_y-_MPa",
            "phi_ctrl_x+_mm", "s_eff_x+_mm", "phi_lim_x+_EC2_mm", "s_lim_x+_EC2_mm",
            "phi_ctrl_x-_mm", "s_eff_x-_mm", "phi_lim_x-_EC2_mm", "s_lim_x-_EC2_mm",
            "phi_ctrl_y+_mm", "s_eff_y+_mm", "phi_lim_y+_EC2_mm", "s_lim_y+_EC2_mm",
            "phi_ctrl_y-_mm", "s_eff_y-_mm", "phi_lim_y-_EC2_mm", "s_lim_y-_EC2_mm",
            "Crack_status_x+", "Crack_status_x-", "Crack_status_y+", "Crack_status_y-",
            "util_M", "util_Qx", "util_Qy", "util_max", "Status", "Verification_reason", "Recommended_action", "Notes"
        ]
        out = results[[c for c in cols if c in results.columns]].copy()
        out["Formula_As"] = "As = MEd/(0,87*fyd*z), z≈0,9d; mínimos EC2 9.6, 9.3.1.1 e 7.3.2 aplicados por face"
        out["Formula_VRdc"] = "VRd,c = [CRdc*k*(100*rho*fck)^(1/3)]*b*d, com vmin; faixa b=1000 mm"
        out["Formula_wk"] = "wk = sr,max*(eps_sm-eps_cm), usando As final real da face traccionada; verificar apenas com ELS quase-permanente"
        return out

    def build_panel_summary(self, results: pd.DataFrame) -> pd.DataFrame:
        if results is None or results.empty:
            return pd.DataFrame()
        rows = []
        for panel, grp in results.groupby("Panel", dropna=False):
            g = grp.copy()
            util_idx = g["util_max"].fillna(-1.0).idxmax()
            m_idx = g["util_M"].fillna(-1.0).idxmax()
            qx_idx = g["util_Qx"].fillna(-1.0).idxmax()
            qy_idx = g["util_Qy"].fillna(-1.0).idxmax()
            r_util = g.loc[util_idx]
            r_m = g.loc[m_idx]
            r_qx = g.loc[qx_idx]
            r_qy = g.loc[qy_idx]

            status_values = set(str(x) for x in g["Status"].dropna())
            if "Não conforme" in status_values:
                status = "Não conforme"
            elif "Verificar" in status_values:
                status = "Verificar"
            elif "Dados insuficientes" in status_values:
                status = "Dados insuficientes"
            elif "OK*" in status_values:
                status = "OK*"
            else:
                status = "OK"
            notes = sorted({str(x) for x in g.get("Notes", pd.Series(dtype=str)).dropna() if str(x).strip()})
            reasons = sorted({str(x) for x in g.get("Verification_reason", pd.Series(dtype=str)).dropna() if str(x).strip()})
            actions = sorted({str(x) for x in g.get("Recommended_action", pd.Series(dtype=str)).dropna() if str(x).strip()})
            rows.append({
                "Panel": panel,
                "n_linhas_analisadas": len(g),
                "Status_global": status,
                "Motivo_estado": " | ".join(reasons[:5]),
                "Acção_recomendada": " | ".join(actions[:5]),
                "util_max": g["util_max"].max(),
                "caso_governante": r_util.get("Case", ""),
                "node_governante": r_util.get("Node", ""),
                "util_M_max": g["util_M"].max(),
                "caso_governante_M": r_m.get("Case", ""),
                "node_governante_M": r_m.get("Node", ""),
                "util_Qx_max": g["util_Qx"].max(),
                "caso_governante_Qx": r_qx.get("Case", ""),
                "node_governante_Qx": r_qx.get("Node", ""),
                "util_Qy_max": g["util_Qy"].max(),
                "util_N_prelim_max": g["util_N_prelim"].max() if "util_N_prelim" in g.columns else 0.0,
                "sigma_N_vertical_max_MPa": g["sigma_N_vertical_MPa"].max() if "sigma_N_vertical_MPa" in g.columns else 0.0,
                "caso_governante_Qy": r_qy.get("Case", ""),
                "node_governante_Qy": r_qy.get("Node", ""),
                "MXX_abs_max_kNm_m": g["MXX_kNm_m"].abs().max(),
                "MYY_abs_max_kNm_m": g["MYY_kNm_m"].abs().max(),
                "MXY_abs_max_kNm_m": g["MXY_kNm_m"].abs().max(),
                "QXX_abs_max_kN_m": g["QXX_kN_m"].abs().max(),
                "QYY_abs_max_kN_m": g["QYY_kN_m"].abs().max(),
                "Asx_req_max_mm2_m": max(g["Asx+_req_mm2_m"].max(), g["Asx-_req_mm2_m"].max()),
                "Asy_req_max_mm2_m": max(g["Asy+_req_mm2_m"].max(), g["Asy-_req_mm2_m"].max()),
                "X+_max_rebar": r_util.get("X+_rebar", ""),
                "X-_max_rebar": r_util.get("X-_rebar", ""),
                "Y+_max_rebar": r_util.get("Y+_rebar", ""),
                "Y-_max_rebar": r_util.get("Y-_rebar", ""),
                "Transverse_links": "Sim" if (g["Transverse_links_required"] == "Sim").any() else "Não",
                "Notas": " | ".join(notes[:4]),
            })
        return pd.DataFrame(rows).sort_values(["Status_global", "Panel"], ascending=[False, True]).reset_index(drop=True)

    def build_governing_design(self, results: pd.DataFrame) -> pd.DataFrame:
        if results is None or results.empty:
            return pd.DataFrame()
        rows = []
        for panel, g in results.groupby("Panel", dropna=False):
            gx = g.copy()
            def row_at_max(col):
                return gx.loc[gx[col].fillna(-1.0).idxmax()]
            rxp = row_at_max("Asx+_req_mm2_m")
            rxn = row_at_max("Asx-_req_mm2_m")
            ryp = row_at_max("Asy+_req_mm2_m")
            ryn = row_at_max("Asy-_req_mm2_m")
            r_util = row_at_max("util_max")
            status_values = set(str(x) for x in gx["Status"].dropna())
            if "Não conforme" in status_values:
                status = "Não conforme"
            elif "Verificar" in status_values:
                status = "Verificar"
            elif "Dados insuficientes" in status_values:
                status = "Dados insuficientes"
            elif "OK*" in status_values:
                status = "OK*"
            else:
                status = "OK"
            reasons = sorted({str(x) for x in gx.get("Verification_reason", pd.Series(dtype=str)).dropna() if str(x).strip()})
            actions = sorted({str(x) for x in gx.get("Recommended_action", pd.Series(dtype=str)).dropna() if str(x).strip()})
            rows.append({
                "Panel": panel,
                "Status": status,
                "Motivo_estado": " | ".join(reasons[:5]),
                "Acção_recomendada": " | ".join(actions[:5]),
                "util_max": gx["util_max"].max(),
                "governing_case": r_util.get("Case", ""),
                "governing_node": r_util.get("Node", ""),
                "X+_adopted": rxp.get("X+_rebar", ""),
                "Asx+_req_max_mm2_m": gx["Asx+_req_mm2_m"].max(),
                "Asx+_prov_mm2_m": rxp.get("Asx+_prov_mm2_m", 0.0),
                "X-_adopted": rxn.get("X-_rebar", ""),
                "Asx-_req_max_mm2_m": gx["Asx-_req_mm2_m"].max(),
                "Asx-_prov_mm2_m": rxn.get("Asx-_prov_mm2_m", 0.0),
                "Y+_adopted": ryp.get("Y+_rebar", ""),
                "Asy+_req_max_mm2_m": gx["Asy+_req_mm2_m"].max(),
                "Asy+_prov_mm2_m": ryp.get("Asy+_prov_mm2_m", 0.0),
                "Y-_adopted": ryn.get("Y-_rebar", ""),
                "Asy-_req_max_mm2_m": gx["Asy-_req_mm2_m"].max(),
                "Asy-_prov_mm2_m": ryn.get("Asy-_prov_mm2_m", 0.0),
                "Transverse_links": "Sim" if (gx["Transverse_links_required"] == "Sim").any() else "Não",
                "Transverse_links_note": " | ".join(sorted(set(str(x) for x in gx["Transverse_links_note"].dropna() if str(x).strip()))[:2]),
                "Design_note": "Armadura indicada por face. X+/Y+ e X-/Y- correspondem aos sinais positivo/negativo dos momentos Wood-Armer."
            })
        return pd.DataFrame(rows).sort_values(["Status", "Panel"], ascending=[False, True]).reset_index(drop=True)

    def build_rebar_zones(self, governing: pd.DataFrame) -> pd.DataFrame:
        if governing is None or governing.empty:
            return pd.DataFrame()
        keys = ["X+_adopted", "X-_adopted", "Y+_adopted", "Y-_adopted", "Transverse_links"]
        rows = []
        for i, (key_vals, grp) in enumerate(governing.groupby(keys, dropna=False), start=1):
            panels = [str(p) for p in grp["Panel"].tolist()]
            worst_util = grp["util_max"].max() if "util_max" in grp.columns else 0.0
            status_values = set(str(x) for x in grp.get("Status", pd.Series(dtype=str)).dropna())
            if "Não conforme" in status_values:
                status = "Não conforme"
            elif "Verificar" in status_values:
                status = "Verificar"
            elif "OK*" in status_values:
                status = "OK*"
            else:
                status = "OK"
            rows.append({
                "Zona": f"Z{i}",
                "Painéis": ", ".join(panels),
                "n_paineis": len(panels),
                "Status": status,
                "util_max": worst_util,
                "X+": key_vals[0],
                "X-": key_vals[1],
                "Y+": key_vals[2],
                "Y-": key_vals[3],
                "Armadura transversal": key_vals[4],
                "Nota": "Agrupamento por solução adoptada; validar continuidade e compatibilização com desenho de armaduras."
            })
        return pd.DataFrame(rows).sort_values(["Status", "Zona"], ascending=[False, True]).reset_index(drop=True)

    def build_rebar_optimization(self, results: pd.DataFrame) -> pd.DataFrame:
        if results is None or results.empty:
            return pd.DataFrame()
        cols = [
            "Panel", "Node", "Case", "Status", "Optimization_mode", "Optimization_iterations",
            "X+_base", "X+_additional", "X+_rebar", "Asx+_req_mm2_m", "Asx+_prov_mm2_m", "X+_optimization",
            "X-_base", "X-_additional", "X-_rebar", "Asx-_req_mm2_m", "Asx-_prov_mm2_m", "X-_optimization",
            "Y+_base", "Y+_additional", "Y+_rebar", "Asy+_req_mm2_m", "Asy+_prov_mm2_m", "Y+_optimization",
            "Y-_base", "Y-_additional", "Y-_rebar", "Asy-_req_mm2_m", "Asy-_prov_mm2_m", "Y-_optimization",
            "util_M", "util_Qx", "util_Qy", "util_max", "Verification_reason", "Recommended_action"
        ]
        out = results[[c for c in cols if c in results.columns]].copy()
        if "util_max" in out.columns:
            out = out.sort_values(["Status", "util_max"], ascending=[False, False])
        return out.reset_index(drop=True)

    def build_normative_notes(self, designer: WallDesigner) -> pd.DataFrame:
        asv_min, ash_min, asv_max = designer.min_wall_reinf_total()
        sx_vert = designer.max_spacing_for_direction(True)
        sx_hor = designer.max_spacing_for_direction(False)
        notes = [
            ("Âmbito", "Painéis/placas", "A tabela com MXX, MYY, MXY, QXX e QYY é tratada como esforços por metro em parede modelada como elemento de placa/casca."),
            ("Flexão", designer.wood_armer_method, "MXY é tratado pelo método seleccionado. No modo conservativo, MXY é convertido em acréscimo absoluto nos momentos de dimensionamento nas duas direcções."),
            ("Eixos locais", "Y vertical" if designer.local_y_is_vertical else "X vertical", "Os resultados dependem da orientação dos eixos locais dos painéis. Confirmar esta convenção antes de adoptar as armaduras."),
            ("Troca X/Y", "Activada" if designer.swap_local_axes else "Desactivada", "Quando activada, MXX/MYY, QXX/QYY e NXX/NYY são trocados antes do cálculo."),
            ("Mínimo vertical", f"Asv,min = {asv_min:.1f} mm²/m", "Adoptado 0,002 Ac por metro de parede, distribuído pelas duas faces."),
            ("Máximo vertical", f"Asv,max = {asv_max:.1f} mm²/m", "Adoptado limite de 0,04 Ac para armadura vertical total por metro."),
            ("Mínimo horizontal", f"Ash,min = {ash_min:.1f} mm²/m", "Adoptado o maior entre 25% da armadura vertical mínima e 0,001 Ac, distribuído pelas duas faces."),
            ("Espaçamento vertical", f"smax = {sx_vert:.0f} mm", "Adoptado min(3t; 400 mm), eventualmente limitado a 250 mm quando o controlo de fendilhação está activo."),
            ("Espaçamento horizontal", f"smax = {sx_hor:.0f} mm", "Adoptado 400 mm, eventualmente limitado a 250 mm quando o controlo de fendilhação está activo."),
            ("Distância livre", "Verificada na escolha automática", "A solução automática rejeita combinações em que a distância livre seja inferior a max(Ø; dg+5; 20 mm)."),
            ("Corte", "VRd,c tipo laje + VRd,max", "A verificação ao esforço transverso é feita sem armadura específica. Se VEd>VRd,c, o programa assinala 'Verificar'; se VEd>VRd,max simplificado, assinala 'Não conforme'."),
            ("Fendilhação", f"wmax = {designer.wmax_mm:.2f} mm", f"Sem selecção de wk: controlo por diâmetro/espaçamento. Com wk seleccionado: cálculo apenas para a combinação quase-permanente indicada ({designer.qp_case or 'não indicada'})."),
            ("Armadura mínima por fendilhação", f"As,min,crack = {designer.crack_min_as_per_face():.1f} mm²/m/face", "Aplicada aproximação de EC2 7.3.2 com Act igual à meia espessura traccionada por face."),
            ("Unidades", f"M: {designer.moment_unit}; Q: {designer.shear_unit}", "Os valores são convertidos internamente para kNm/m e kN/m. Confirmar sempre as unidades da tabela de origem antes de adoptar os resultados."),
            ("Optimização de armaduras", "Base + reforços" if designer.optimize_rebar else "Automática directa", f"Base vertical Ø{designer.base_vertical_phi:.0f}//{designer.base_vertical_spacing:.0f}; base horizontal Ø{designer.base_horizontal_phi:.0f}//{designer.base_horizontal_spacing:.0f}. Quando a base é insuficiente, o programa tenta adicionar reforço local antes de declarar insuficiência à flexão ou corte."),
            ("Limitação", "Sem verificação completa N-M", "A verificação de parede comprimida, estabilidade, segunda ordem e flexão composta requer esforços normais/membrana e geometria global da parede."),
        ]
        return pd.DataFrame(notes, columns=["Tema", "Critério adoptado", "Nota"])

    def run_design(self):
        err = self.validate_inputs()
        if err:
            messagebox.showwarning("Aviso", err)
            return
        designer = WallDesigner(
            thickness_mm=safe_float(self.var_thickness.get(), 200.0),
            cover_mm=safe_float(self.var_cover.get(), 35.0),
            concrete_class=self.var_concrete.get(),
            fyk=safe_float(self.var_steel.get(), 500.0),
            local_y_is_vertical=self.var_y_vertical.get(),
            crack_spacing_limit_mm=250.0 if self.var_crack_spacing.get() else None,
            wood_armer_method=self.var_wa_method.get(),
            swap_local_axes=self.var_swap_axes.get(),
            crack_check_enabled=True,
            wk_check_enabled=self.var_crack_check.get(),
            qp_case=self.var_qp_case.get(),
            moment_unit=self.var_moment_unit.get(),
            shear_unit=self.var_shear_unit.get(),
            combo_type=self.var_combo_type.get(),
            wmax_mm=safe_float(self.var_wmax.get(), 0.30),
            phi_min_mm=safe_float(self.var_phi_min.get(), 8.0),
            phi_max_mm=safe_float(self.var_phi_max.get(), 16.0),
            optimize_rebar=self.var_optimize_rebar.get(),
            rebar_strategy=self.var_rebar_strategy.get(),
            base_vertical_phi=safe_float(self.var_base_v_phi.get(), 10.0),
            base_vertical_spacing=safe_float(self.var_base_v_spacing.get(), 200.0),
            base_horizontal_phi=safe_float(self.var_base_h_phi.get(), 8.0),
            base_horizontal_spacing=safe_float(self.var_base_h_spacing.get(), 200.0),
        )
        if designer.wk_check_enabled and not designer.qp_case:
            messagebox.showwarning("Aviso", "Foi seleccionada a verificação wk, mas não foi indicado o número da combinação quase-permanente.")
            return
        if self.var_reduce.get():
            df_gov = reduce_to_governing_cases(self.df_clean)
            if designer.wk_check_enabled and designer.qp_case and "case" in self.df_clean.columns:
                mask_qp = self.df_clean["case"].map(designer.is_qp_case)
                df_qp = self.df_clean.loc[mask_qp].copy()
                df_in = pd.concat([df_gov, df_qp], ignore_index=True).drop_duplicates(subset=["__row_order"], keep="first")
            else:
                df_in = df_gov
        else:
            df_in = self.df_clean.copy()
        try:
            self.progress_var.set(0.0)
            self.status_var.set("A calcular...")
            self.update_idletasks()
            rows = []
            total = max(len(df_in), 1)
            for i, (_, r) in enumerate(df_in.iterrows(), start=1):
                rows.append(designer.design_one(r))
                if i == 1 or i == total or i % max(1, total // 100) == 0:
                    self.progress_var.set(100.0 * i / total)
                    self.status_var.set(f"A calcular... {i}/{total} linhas")
                    self.update_idletasks()
            self.df_results = pd.DataFrame(rows)
            self.status_var.set("A preparar resumos...")
            self.update_idletasks()
            self.df_summary = self.build_panel_summary(self.df_results)
            self.df_governing = self.build_governing_design(self.df_results)
            self.df_zones = self.build_rebar_zones(self.df_governing)
            self.df_optimization = self.build_rebar_optimization(self.df_results)
            self.df_unit_check = self.build_unit_check(df_in, designer)
            self.df_data_validation = self.build_data_validation(df_in, designer)
            self.df_diagnostic = self.build_diagnostic(self.df_results, self.df_data_validation)
            self.df_sketches = self.build_rebar_sketches(self.df_zones)
            self.df_detailed = self.build_detailed_check(self.df_results)
            self.df_notes = self.build_normative_notes(designer)
            self.show_df(self.tree_results, self.df_results)
            self.show_df(self.tree_summary, self.df_summary)
            self.show_df(self.tree_governing, self.df_governing)
            self.show_df(self.tree_zones, self.df_zones)
            self.show_df(self.tree_optimization, self.df_optimization)
            self.show_df(self.tree_diagnostic, self.df_diagnostic)
            self.show_df(self.tree_validation, self.df_data_validation)
            self.show_df(self.tree_sketches, self.df_sketches)
            self.show_df(self.tree_unit, self.df_unit_check)
            self.show_df(self.tree_detailed, self.df_detailed)
            self.show_df(self.tree_notes, self.df_notes)
            qp_msg = ""
            if designer.wk_check_enabled:
                n_qp = int((self.df_results.get("is_QP_crack_case", pd.Series(dtype=str)) == "Sim").sum())
                qp_msg = f"; wk QP: {n_qp} linhas"
                if n_qp == 0:
                    messagebox.showwarning("Aviso", f"Não foram encontradas linhas para a combinação quase-permanente indicada: {designer.qp_case}")
            self.progress_var.set(100.0)
            self.status_var.set(f"Cálculo concluído: {len(self.df_results)} linhas analisadas; {len(self.df_summary)} painéis resumidos{qp_msg}.")
        except Exception as e:
            self.progress_var.set(0.0)
            messagebox.showerror("Erro", str(e))
            self.status_var.set("Falha no cálculo.")

    def export_excel(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar resultados",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return

        # Garante extensão correcta. Evita ficheiros com extensão inválida,
        # por exemplo .xls ou nomes sem extensão que depois o Excel recusa abrir.
        import os
        root, ext = os.path.splitext(path)
        if ext.lower() != ".xlsx":
            path = root + ".xlsx" if ext else path + ".xlsx"

        def safe_sheet_name(name: str) -> str:
            bad = r'[]:*?/\\'
            out = "".join("_" if ch in bad else ch for ch in str(name))[:31]
            return out or "Sheet"

        def has_df(df):
            return df is not None and isinstance(df, pd.DataFrame) and not df.empty

        def pick_cols(df, cols):
            if not has_df(df):
                return pd.DataFrame()
            present = [c for c in cols if c in df.columns]
            if not present:
                return pd.DataFrame()
            return df[present].copy()

        try:
            self.progress_var.set(0.0)
            self.status_var.set("A preparar exportação Excel...")
            self.update_idletasks()

            # ---- folhas estruturadas, menos densas ----
            export_dt = datetime.now()
            metadata = pd.DataFrame([
                ["Programa", APP_NAME],
                ["Versão", APP_VERSION],
                ["Autor", APP_AUTHOR],
                ["Repositório", GITHUB_URL],
                ["Data de exportação", export_dt.strftime("%Y-%m-%d %H:%M")],
                ["Tipo de ficheiro", "Resultados de cálculo"],
                ["Norma de referência", "Eurocódigo 2 / NP EN 1992-1-1"],
                ["Âmbito", "Paredes modeladas como painéis/placas, com esforços por metro"],
                ["Limitações", "Não inclui compressão global, flexão composta, estabilidade e efeitos de segunda ordem"],
                ["Unidades internas", "Momentos em kNm/m; corte em kN/m"],
                ["Descrição", APP_XLSX_DESCRIPTION],
            ], columns=["Campo", "Valor"])

            input_params = pd.DataFrame([
                ["Espessura [mm]", self.var_thickness.get()],
                ["Recobrimento [mm]", self.var_cover.get()],
                ["Betão", self.var_concrete.get()],
                ["Aço fyk [MPa]", self.var_steel.get()],
                ["Eixo local Y = vertical", "Sim" if self.var_y_vertical.get() else "Não"],
                ["Trocar eixos X ↔ Y", "Sim" if self.var_swap_axes.get() else "Não"],
                ["Método MXY", self.var_wa_method.get()],
                ["Verificar wk", "Sim" if self.var_crack_check.get() else "Não"],
                ["Combinação quase-permanente", self.var_qp_case.get()],
                ["wmax [mm]", self.var_wmax.get()],
                ["Unidade dos momentos", self.var_moment_unit.get()],
                ["Unidade do corte", self.var_shear_unit.get()],
                ["Ø min / Ø max [mm]", f"{self.var_phi_min.get()} / {self.var_phi_max.get()}"],
                ["Optimizar armadura", "Sim" if self.var_optimize_rebar.get() else "Não"],
                ["Estratégia", self.var_rebar_strategy.get()],
                ["Base vertical", f"Ø{self.var_base_v_phi.get()}//{self.var_base_v_spacing.get()}"],
                ["Base horizontal", f"Ø{self.var_base_h_phi.get()}//{self.var_base_h_spacing.get()}"],
            ], columns=["Parâmetro", "Valor"] )

            readme = pd.DataFrame([
                ["WallsEC2", "Exportação organizada de resultados de paredes de betão armado"],
                ["Formato", "Workbook .xlsx válido, com resultados separados por tema"],
                ["Linhas analisadas", len(self.df_results) if has_df(self.df_results) else 0],
                ["Painéis resumidos", len(self.df_summary) if has_df(self.df_summary) else 0],
                ["Folhas principais", "Resumo por painel, armaduras adoptadas, zonas, diagnóstico, validação, resultados por tema"],
                ["Nota", "A folha '15_Full_Results' mantém a tabela completa para auditoria; as folhas 09 a 13 separam os resultados por assunto."],
            ], columns=["Item", "Descrição"] )

            row_core = pick_cols(self.df_results, [
                "Panel", "Node", "Case", "is_QP_crack_case", "Status", "Verification_reason", "Recommended_action",
                "util_max", "util_M", "util_Qx", "util_Qy", "wk_max_mm"
            ])
            row_flexure = pick_cols(self.df_results, [
                "Panel", "Node", "Case", "MXX_kNm_m", "MYY_kNm_m", "MXY_kNm_m", "WA_method", "MXY_ratio",
                "Mx+_WA", "Mx-_WA", "My+_WA", "My-_WA",
                "Asx+_req_mm2_m", "Asx-_req_mm2_m", "Asy+_req_mm2_m", "Asy-_req_mm2_m",
                "MRd_x+_kNm_m", "MRd_x-_kNm_m", "MRd_y+_kNm_m", "MRd_y-_kNm_m", "util_M"
            ])
            row_shear = pick_cols(self.df_results, [
                "Panel", "Node", "Case", "QXX_kN_m", "QYY_kN_m", "VRdc_x_kN_m", "VRdc_y_kN_m",
                "VRdmax_x_kN_m", "VRdmax_y_kN_m", "util_Qx", "util_Qy",
                "util_Qx_VRdmax", "util_Qy_VRdmax", "Transverse_links_required", "Transverse_links_note"
            ])
            row_cracking = pick_cols(self.df_results, [
                "Panel", "Node", "Case", "is_QP_crack_case", "wmax_mm",
                "wk_x+_mm", "wk_x-_mm", "wk_y+_mm", "wk_y-_mm", "wk_max_mm",
                "sigma_s_x+_MPa", "sigma_s_x-_MPa", "sigma_s_y+_MPa", "sigma_s_y-_MPa",
                "phi_ctrl_x+_mm", "phi_ctrl_x-_mm", "phi_ctrl_y+_mm", "phi_ctrl_y-_mm",
                "s_eff_x+_mm", "s_eff_x-_mm", "s_eff_y+_mm", "s_eff_y-_mm",
                "phi_lim_x+_EC2_mm", "phi_lim_x-_EC2_mm", "phi_lim_y+_EC2_mm", "phi_lim_y-_EC2_mm",
                "s_lim_x+_EC2_mm", "s_lim_x-_EC2_mm", "s_lim_y+_EC2_mm", "s_lim_y-_EC2_mm",
                "Crack_status_x+", "Crack_status_x-", "Crack_status_y+", "Crack_status_y-",
                "Crack_reason_x+", "Crack_reason_x-", "Crack_reason_y+", "Crack_reason_y-",
                "Crack_control_status", "Crack_control_note"
            ])
            row_rebar = pick_cols(self.df_results, [
                "Panel", "Node", "Case",
                "X+_rebar", "X-_rebar", "Y+_rebar", "Y-_rebar",
                "X+_base", "X-_base", "Y+_base", "Y-_base",
                "X+_additional", "X-_additional", "Y+_additional", "Y-_additional",
                "Asx+_prov_mm2_m", "Asx-_prov_mm2_m", "Asy+_prov_mm2_m", "Asy-_prov_mm2_m",
                "sX_max_mm", "sY_max_mm", "Optimization_mode", "Optimization_iterations"
            ])

            sheets = [
                ("00_METADATA", metadata),
                ("00_README", readme),
                ("01_Input_Data", input_params),
                ("02_Panel_Summary", self.df_summary),
                ("03_Adopted_Rebar", self.df_governing),
                ("04_Rebar_Zones", self.df_zones),
                ("05_Optimization", self.df_optimization),
                ("06_Diagnostic", self.df_diagnostic),
                ("07_Data_Validation", self.df_data_validation),
                ("08_Unit_Check", self.df_unit_check),
                ("09_Row_Core", row_core),
                ("10_Row_Flexure", row_flexure),
                ("11_Row_Shear", row_shear),
                ("12_Row_Cracking", row_cracking),
                ("13_Row_Rebar", row_rebar),
                ("14_Detailed_Check", self.df_detailed),
                ("15_Full_Results", self.df_results),
                ("16_Raw_Table", self.df_clean),
                ("17_EC2_Notes", self.df_notes),
            ]

            self.progress_var.set(15.0)
            self.status_var.set("A escrever folhas Excel...")
            self.update_idletasks()

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                for i, (name, df) in enumerate(sheets):
                    df_to_write = df if has_df(df) else pd.DataFrame({"Nota": ["Sem dados disponíveis."]})
                    df_to_write.to_excel(writer, sheet_name=safe_sheet_name(name), index=False)
                    self.progress_var.set(15.0 + 45.0 * (i + 1) / max(1, len(sheets)))
                    self.status_var.set(f"A escrever folha {i+1}/{len(sheets)}: {name}")
                    self.update_idletasks()

                # ---- formatação profissional do workbook ----
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                from openpyxl.formatting.rule import FormulaRule

                wb = writer.book
                props = wb.properties
                props.title = f"{APP_NAME} - Resultados de Cálculo"
                props.subject = APP_SUBJECT
                props.creator = APP_AUTHOR
                props.lastModifiedBy = APP_NAME
                props.description = APP_XLSX_DESCRIPTION
                props.keywords = APP_KEYWORDS
                props.category = APP_CATEGORY
                props.created = export_dt
                props.modified = export_dt

                header_fill = PatternFill("solid", fgColor="1F4E78")
                header_font = Font(color="FFFFFF", bold=True)
                title_fill = PatternFill("solid", fgColor="D9EAF7")
                light_fill = PatternFill("solid", fgColor="F7F9FB")
                thin = Side(style="thin", color="D9E2EC")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for ws in wb.worksheets:
                    ws.freeze_panes = "A2"
                    ws.sheet_view.showGridLines = False
                    max_row = ws.max_row
                    max_col = ws.max_column
                    if max_row >= 1 and max_col >= 1:
                        ws.auto_filter.ref = ws.dimensions
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = border
                        ws.row_dimensions[1].height = 24
                        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                            for cell in row:
                                cell.border = border
                                cell.alignment = Alignment(vertical="top", wrap_text=True)
                                if isinstance(cell.value, float):
                                    cell.number_format = "0.000"
                        # zebra leve para legibilidade
                        for r in range(2, max_row + 1):
                            if r % 2 == 0:
                                for c in range(1, max_col + 1):
                                    ws.cell(r, c).fill = light_fill

                    # larguras de colunas com limites para evitar folhas ilegíveis
                    for col_idx in range(1, max_col + 1):
                        letter = get_column_letter(col_idx)
                        header = str(ws.cell(1, col_idx).value or "")
                        sample_vals = [str(ws.cell(r, col_idx).value or "") for r in range(1, min(max_row, 80) + 1)]
                        width = max([len(header)] + [min(len(v), 50) for v in sample_vals]) + 2
                        if any(k in header.lower() for k in ["note", "nota", "reason", "action", "problema", "descrição", "descricao", "croquis"]):
                            width = min(max(width, 22), 46)
                        elif any(k in header.lower() for k in ["status", "case", "panel", "node"]):
                            width = min(max(width, 10), 18)
                        else:
                            width = min(max(width, 10), 22)
                        ws.column_dimensions[letter].width = width

                    # formatação especial por folha
                    if ws.title in {"00_METADATA", "00_README", "01_Input_Data"}:
                        ws.column_dimensions["A"].width = 28
                        ws.column_dimensions["B"].width = 80
                        for row in ws.iter_rows(min_row=2, max_row=max_row):
                            if row and row[0].value:
                                row[0].font = Font(bold=True, color="1F4E78")

                    # Condicionais em colunas de estado/status
                    headers = {str(ws.cell(1, c).value): c for c in range(1, max_col + 1)}
                    for status_header in ["Status", "Status_global", "Estado", "Crack_control_status"]:
                        if status_header in headers and max_row > 1:
                            col = get_column_letter(headers[status_header])
                            rng = f"{col}2:{col}{max_row}"
                            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("Não conforme",{col}2))'], fill=PatternFill("solid", fgColor="F4CCCC")))
                            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("Verificar",{col}2))'], fill=PatternFill("solid", fgColor="FFF2CC")))
                            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("OK",{col}2))'], fill=PatternFill("solid", fgColor="D9EAD3")))

                # Larguras específicas em folhas densas
                for dense in ["15_Full_Results", "16_Raw_Table"]:
                    if dense in wb.sheetnames:
                        ws = wb[dense]
                        ws.freeze_panes = "A2"
                        # sheet muito densa fica no fim e serve apenas para auditoria
                        for col_idx in range(1, ws.max_column + 1):
                            ws.column_dimensions[get_column_letter(col_idx)].width = 14

                # Índice visual na primeira folha
                if "00_README" in wb.sheetnames:
                    ws = wb["00_README"]
                    start = ws.max_row + 3
                    ws.cell(start, 1, "Índice de folhas").font = Font(bold=True, color="1F4E78")
                    ws.cell(start, 1).fill = title_fill
                    ws.cell(start, 2).fill = title_fill
                    ws.cell(start, 1).border = border
                    ws.cell(start, 2).border = border
                    for j, sheet_name in enumerate(wb.sheetnames, start=start + 1):
                        ws.cell(j, 1, sheet_name)
                        ws.cell(j, 2, self._excel_sheet_description(sheet_name) if hasattr(self, "_excel_sheet_description") else "")
                        ws.cell(j, 1).border = border
                        ws.cell(j, 2).border = border
                    ws.auto_filter.ref = f"A1:B{ws.max_row}"

            self.progress_var.set(100.0)
            self.status_var.set(f"Excel organizado exportado: {path}")
            messagebox.showinfo("Exportação concluída", f"Excel exportado com sucesso:\n{path}")

        except PermissionError:
            self.progress_var.set(0.0)
            self.status_var.set("Erro ao exportar Excel: ficheiro aberto ou sem permissão.")
            messagebox.showerror(
                "Erro",
                "Não foi possível gravar o ficheiro. Feche o Excel se o ficheiro estiver aberto e tente novamente."
            )
        except Exception as e:
            self.progress_var.set(0.0)
            self.status_var.set("Erro ao exportar Excel.")
            messagebox.showerror("Erro", str(e))

    def _excel_sheet_description(self, sheet_name: str) -> str:
        descriptions = {
            "00_METADATA": "Metadados do ficheiro e identificação da exportação",
            "00_README": "Resumo da exportação e índice do workbook",
            "01_Input_Data": "Parâmetros usados no cálculo",
            "02_Panel_Summary": "Resumo governante por painel",
            "03_Adopted_Rebar": "Armadura adoptada por painel",
            "04_Rebar_Zones": "Agrupamento de painéis por zonas de armadura",
            "05_Optimization": "Histórico/resultado da optimização de armaduras",
            "06_Diagnostic": "Avisos e acções recomendadas",
            "07_Data_Validation": "Validação da tabela importada",
            "08_Unit_Check": "Unidades adoptadas e factores de conversão",
            "09_Row_Core": "Resultados essenciais por linha/caso",
            "10_Row_Flexure": "Resultados de flexão e armadura requerida",
            "11_Row_Shear": "Verificação ao esforço transverso",
            "12_Row_Cracking": "Verificação de fendilhação / wk",
            "13_Row_Rebar": "Armadura base, adicional e final por linha",
            "14_Detailed_Check": "Cálculo detalhado de linhas seleccionadas/críticas",
            "15_Full_Results": "Tabela completa de resultados para auditoria",
            "16_Raw_Table": "Tabela importada já normalizada",
            "17_EC2_Notes": "Notas normativas e limitações",
        }
        return descriptions.get(sheet_name, "")

    def export_pdf_report(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar relatório PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
        )
        if not path:
            return
        try:
            from datetime import datetime
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, NextPageTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether

            self.progress_var.set(0.0)
            self.status_var.set("A gerar relatório PDF...")
            self.update_idletasks()

            page_size = landscape(A4)
            margin_l = 16 * mm
            margin_r = 16 * mm
            margin_t = 18 * mm
            margin_b = 15 * mm
            doc = BaseDocTemplate(
                path,
                pagesize=page_size,
                rightMargin=margin_r,
                leftMargin=margin_l,
                topMargin=margin_t,
                bottomMargin=margin_b,
            )
            styles = getSampleStyleSheet()
            # Formatação pedida para a memória de ccálculo:
            # texto em Courier 10, subtítulos em Courier-Bold 12, espaçamento 1.5.
            styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
            styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
            styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=10, leading=15, spaceAfter=6))
            styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=8, leading=12))
            styles.add(ParagraphStyle(name="Cell", parent=styles["Small"], alignment=TA_LEFT, fontName="Courier", fontSize=7, leading=10.5))
            styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=12, leading=18, spaceBefore=10, spaceAfter=20))

            def header_footer(canvas, doc_obj):
                canvas.saveState()
                canvas.setTitle(f"{APP_NAME} - Relatório de Cálculo")
                canvas.setAuthor(APP_AUTHOR)
                canvas.setSubject(APP_SUBJECT)
                canvas.setCreator(f"{APP_NAME} {APP_VERSION}")
                canvas.setKeywords(APP_KEYWORDS)
                try:
                    canvas._doc.info.producer = f"{APP_NAME} / ReportLab"
                except Exception:
                    pass
                w, h = canvas._pagesize
                canvas.setFont("Courier", 8)
                canvas.setFillColor(colors.grey)
                header_name = "WallsEC2"
                header_txt = header_name + " | Dimensionamento de Pareces (EC2)"
                canvas.drawString(margin_l, h - 10 * mm, header_txt)
                # Link invisível sobre o nome do programa no cabeçalho.
                header_w = canvas.stringWidth(header_name, "Courier", 8)
                canvas.linkURL(GITHUB_URL, (margin_l, h - 11.5 * mm, margin_l + header_w, h - 8.5 * mm), relative=0)
                canvas.drawRightString(w - margin_r, h - 10 * mm, f"Página {doc_obj.page}")
                canvas.line(margin_l, h - 12 * mm, w - margin_r, h - 12 * mm)
                footer_name = "WallsEC2"
                footer_txt = footer_name + " | " + datetime.now().strftime("%Y-%m-%d %H:%M")
                footer_x = w / 2.0
                canvas.drawCentredString(footer_x, 8 * mm, footer_txt)
                # Link invisível sobre o nome do programa no rodapé.
                footer_name_w = canvas.stringWidth(footer_name, "Courier", 8)
                footer_total_w = canvas.stringWidth(footer_txt, "Courier", 8)
                x0 = footer_x - footer_total_w / 2.0
                canvas.linkURL(GITHUB_URL, (x0, 6.5 * mm, x0 + footer_name_w, 9.5 * mm), relative=0)
                canvas.restoreState()

            # Template único: todo o relatório, incluindo o apêndice, em A4 paisagem.
            frame_a4 = Frame(margin_l, margin_b, page_size[0] - margin_l - margin_r, page_size[1] - margin_t - margin_b, id="A4L_frame")
            doc.addPageTemplates([
                PageTemplate(id="A4L", pagesize=page_size, frames=[frame_a4], onPage=header_footer),
            ])

            def fmt(v):
                if isinstance(v, float):
                    if not math.isfinite(v):
                        return "-"
                    return f"{v:.3f}"
                return "" if pd.isna(v) else str(v)

            def escape_pdf_text(s: str) -> str:
                return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

            def simplify_status_reason(value) -> str:
                """Simplifica motivos extensos apenas para apresentação no PDF."""
                s = fmt(value).strip()
                sl = s.lower()
                if not s or s in {"-", "nan"}:
                    return "-"
                if "v ed" in sl or "ved" in sl or "vrd" in sl or "corte" in sl:
                    return "Corte"
                if "mxy" in sl:
                    return "MXY dominante"
                if "wk" in sl or "fend" in sl:
                    if "limite" in sl:
                        return "Fendilhação no limite"
                    return "Fendilhação"
                if "unid" in sl or "unit" in sl:
                    return "Confirmar unidades"
                if "eixo" in sl or "axis" in sl:
                    return "Confirmar eixos"
                if "base" in sl and "insuf" in sl:
                    return "Base insuficiente"
                if len(s) > 38:
                    return s[:35].rstrip() + "..."
                return s

            def simplify_recommendation(value) -> str:
                """Gera uma nota curta para o PDF. O detalhe completo fica no Excel."""
                s = fmt(value).strip()
                sl = s.lower()
                if not s or s in {"-", "nan"}:
                    return "-"
                if "unid" in sl or "unit" in sl:
                    return "Confirmar unidades"
                if "eixo" in sl or "axis" in sl:
                    return "Confirmar eixos locais"
                if "fend" in sl or "wk" in sl or "els" in sl:
                    return "Confirmar ELS/fendilhação"
                if "corte" in sl or "vrd" in sl or "ved" in sl:
                    return "Rever corte"
                if "espess" in sl or "geometr" in sl:
                    return "Rever geometria"
                if len(s) > 36:
                    return s[:33].rstrip() + "..."
                return s

            PDF_HEADER_MAP = {
                "Panel": "Painel",
                "Node": "Nó",
                "Case": "Caso",
                "Status_global": "Estado",
                "Status": "Estado",
                "Motivo_estado": "Motivo",
                "Acção_recomendada": "Nota",
                "Acao_recomendada": "Nota",
                "Recommended_action": "Nota",
                "Verification_reason": "Motivo",
                "caso_governante": "Caso gov.",
                "node_governante": "Nó gov.",
                "util_max": "Util.",
                "util_M_max": "Util. M",
                "util_Qx_max": "Util. Qx",
                "util_Qy_max": "Util. Qy",
                "Transverse_links": "Lig. faces",
                "governing_case": "Caso gov.",
                "is_QP_crack_case": "QP",
                "MXX_kNm_m": "MXX",
                "MYY_kNm_m": "MYY",
                "MXY_kNm_m": "MXY",
                "QXX_kN_m": "QXX",
                "QYY_kN_m": "QYY",
                "wk_max_mm": "wk",
                "Optimization_mode": "Modo",
                "Acção recomendada": "Nota",
                "Acao recomendada": "Nota",
            }

            def display_value(column, value):
                if column == "Motivo_estado" or column == "Verification_reason":
                    return simplify_status_reason(value)
                if column in {"Acção_recomendada", "Acao_recomendada", "Recommended_action", "Acção recomendada", "Acao recomendada"}:
                    return simplify_recommendation(value)
                return fmt(value)

            def as_par(v, column=None):
                value = display_value(column, v)
                # Permite hyperlinks ReportLab já formatados, mantendo o URL invisível no texto.
                if isinstance(value, str) and ("<a href=" in value.lower() or "<link href=" in value.lower()):
                    return Paragraph(value, styles["Cell"])
                return Paragraph(escape_pdf_text(value), styles["Cell"])

            def add_table(title, df, cols=None, max_rows=30, widths=None, note_cols=None, note_title="Notas"):
                story.append(Paragraph(title, styles["Section"]))
                if df is None or df.empty:
                    story.append(Paragraph("Sem dados disponíveis.", styles["Small"]))
                    story.append(Spacer(1, 5))
                    return
                show = df.head(max_rows).copy()
                note_cols = note_cols or []
                if cols is None:
                    cols_use = list(show.columns)
                else:
                    cols_use = [c for c in cols if c in show.columns]
                # As colunas definidas em note_cols não entram no quadro principal; são resumidas abaixo.
                main_cols = [c for c in cols_use if c not in note_cols]
                if not main_cols:
                    story.append(Paragraph("Sem colunas disponíveis para este quadro.", styles["Small"]))
                    return
                data = [[Paragraph(escape_pdf_text(PDF_HEADER_MAP.get(c, c)), styles["Cell"]) for c in main_cols]]
                for _, row in show[main_cols].iterrows():
                    data.append([as_par(row[c], c) for c in main_cols])
                if widths is None:
                    widths = [max(18 * mm, min(34 * mm, 6 * mm + len(str(PDF_HEADER_MAP.get(c, c))) * 1.1 * mm)) for c in main_cols]
                tbl = Table(data, repeatRows=1, colWidths=widths)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7B7B7")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
                ]))
                story.append(tbl)

                # Notas compactas fora do quadro, para evitar tabelas muito largas.
                notes = []
                for col in note_cols:
                    if col in show.columns:
                        for _, row in show.iterrows():
                            txt = simplify_recommendation(row[col])
                            if txt and txt != "-" and txt not in notes:
                                notes.append(txt)
                if notes:
                    story.append(Paragraph(f"{note_title}: " + escape_pdf_text("; ".join(notes[:6])) + ".", styles["Small"]))
                if len(df) > max_rows:
                    story.append(Paragraph(f"Nota: quadro truncado no relatório. Linhas apresentadas: {max_rows} de {len(df)}. Consultar ficheiro Excel para resultados completos.", styles["Small"]))
                story.append(Spacer(1, 6))

            story = []
            story.append(Spacer(1, 18))
            story.append(Paragraph("RELATÓRIO DE CÁLCULO", styles["ReportTitle"]))
            story.append(Paragraph("Dimensionamento de paredes de betão armado segundo EC2", styles["ReportSubtitle"]))
            story.append(Spacer(1, 10))

            project_data = [
                ["Programa", f'<link href="{GITHUB_URL}">WallsEC2</link> - Paredes de Betão Armado (EC2)'],
                ["Data de emissão", datetime.now().strftime("%Y-%m-%d %H:%M")],
                ["Espessura da parede", f"{self.var_thickness.get()} mm"],
                ["Recobrimento nominal", f"{self.var_cover.get()} mm"],
                ["Classe de betão", self.var_concrete.get()],
                ["Classe de Aço", f"fyk = {self.var_steel.get()} MPa"],
                ["Esforços", self.var_wa_method.get()],
                ["Eixos locais", "Y local vertical" if self.var_y_vertical.get() else "X local vertical"],
                ["Troca X/Y", "Sim" if self.var_swap_axes.get() else "Não"],
                ["Unidades de momentos", self.var_moment_unit.get()],
                ["Unidades de esforço transverso", self.var_shear_unit.get()],
                ["Verificação wk", "Sim" if self.var_crack_check.get() else "Não"],
                ["Combinação quase-permanente", self.var_qp_case.get() or "não indicada"],
                ["Limite wmax", f"{self.var_wmax.get()} mm"],
                ["Optimização de armaduras", (self.var_rebar_strategy.get() if self.var_optimize_rebar.get() else "Automática directa")],
                ["Base vertical", f"Ø{self.var_base_v_phi.get()}//{self.var_base_v_spacing.get()}"],
                ["Base horizontal", f"Ø{self.var_base_h_phi.get()}//{self.var_base_h_spacing.get()}"],
            ]
            tbl = Table([[Paragraph("Parâmetro", styles["Cell"]), Paragraph("Valor", styles["Cell"])] ] + [[as_par(a), as_par(b)] for a,b in project_data], colWidths=[60*mm, 110*mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#D9EAF7")),
                ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#B7B7B7")),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F7F7F7")]),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 10))

            n_panels = len(self.df_summary) if self.df_summary is not None else 0
            n_lines = len(self.df_results) if self.df_results is not None else 0
            status_counts = self.df_summary["Status_global"].value_counts().to_dict() if self.df_summary is not None and "Status_global" in self.df_summary.columns else {}
            resumo_txt = (
                f"Foram analisadas {n_lines} linhas de esforços, correspondentes a {n_panels} painéis. "
                f"Estados globais por painel: {status_counts}. "
                "O cálculo considera os painéis como elementos de placa/casca, com dimensionamento por faixas de 1 m. "
                "A verificação à compressão, flexão composta global, estabilidade e efeitos de segunda ordem não está incluída neste anexo."
            )
            story.append(Paragraph("Resumo executivo", styles["Section"]))
            story.append(Paragraph(resumo_txt, styles["BodyCourier"]))
            story.append(Spacer(1, 8))
            self.progress_var.set(20.0); self.update_idletasks()

            add_table("1. Resumo por painel", self.df_summary, cols=[
                "Panel", "Status_global", "Motivo_estado", "Acção_recomendada", "util_max", "caso_governante", "node_governante", "util_M_max", "util_Qx_max", "util_Qy_max", "Transverse_links"
            ], max_rows=36, note_cols=["Acção_recomendada"], widths=[20*mm, 20*mm, 38*mm, 18*mm, 26*mm, 20*mm, 18*mm, 18*mm, 18*mm, 22*mm])
            self.progress_var.set(40.0); self.update_idletasks()

            add_table("2. Armadura adoptada por painel", self.df_governing, cols=[
                "Panel", "Status", "Motivo_estado", "Acção_recomendada", "util_max", "X+_adopted", "X-_adopted", "Y+_adopted", "Y-_adopted", "Transverse_links", "governing_case"
            ], max_rows=36, note_cols=["Acção_recomendada"], widths=[18*mm, 18*mm, 34*mm, 18*mm, 30*mm, 30*mm, 30*mm, 30*mm, 22*mm, 22*mm])
            self.progress_var.set(55.0); self.update_idletasks()

            add_table("3. Zonas de armadura", self.df_zones, cols=[
                "Zona", "Painéis", "Status", "util_max", "X+", "X-", "Y+", "Y-", "Armadura transversal", "Nota"
            ], max_rows=40)

            add_table("4. Optimização de armaduras", self.df_optimization, cols=[
                "Panel", "Status", "Optimization_mode", "X+_base", "X+_additional", "X+_rebar", "Y+_base", "Y+_additional", "Y+_rebar", "util_max"
            ], max_rows=36)
            add_table("5. Diagnóstico", self.df_diagnostic, cols=[
                "Painel", "Severidade", "Problema", "Acção recomendada", "Origem"
            ], max_rows=40, widths=[24*mm, 24*mm, 92*mm, 34*mm, 36*mm])
            story.append(PageBreak())
            self.progress_var.set(65.0); self.update_idletasks()

            add_table("6. Validação da tabela importada", self.df_data_validation, cols=[
                "Categoria", "Item", "Estado", "Resultado", "Nota"
            ], max_rows=50, widths=[35*mm, 45*mm, 28*mm, 50*mm, 95*mm])
            add_table("7. Croquis textual por zona", self.df_sketches, cols=[
                "Zona", "Painéis", "Croquis_textual", "Nota_desenho"
            ], max_rows=20, widths=[18*mm, 50*mm, 125*mm, 72*mm])

            add_table("8. Verificação de unidades e consistência", self.df_unit_check, cols=[
                "Grandeza", "Unidade selecionada", "Factor aplicado", "Máximo convertido", "Unidade interna", "Nota"
            ], max_rows=10)
            add_table("9. Notas normativas e hipóteses de cálculo", self.df_notes, cols=[
                "Tema", "Critério adoptado", "Nota"
            ], max_rows=30, widths=[40*mm, 55*mm, 155*mm])
            self.progress_var.set(80.0); self.update_idletasks()

            # Apêndice reduzido com linhas críticas em A4 paisagem, com colunas compactas.
            if self.df_results is not None and not self.df_results.empty:
                story.append(PageBreak())
                crit = self.df_results.sort_values("util_max", ascending=False).head(20)
                add_table("10. Apêndice - linhas críticas por utilização", crit, cols=[
                    "Panel", "Node", "Case", "is_QP_crack_case", "MXX_kNm_m", "MYY_kNm_m", "MXY_kNm_m", "QXX_kN_m", "QYY_kN_m", "wk_max_mm", "Status", "Recommended_action"
                ], max_rows=20, note_cols=["Recommended_action"], widths=[15*mm, 16*mm, 22*mm, 14*mm, 21*mm, 21*mm, 21*mm, 21*mm, 21*mm, 16*mm, 18*mm])
            self.progress_var.set(90.0); self.update_idletasks()

            doc.build(story)
            self.progress_var.set(100.0)
            self.status_var.set(f"Relatório PDF exportado: {path}")
        except ImportError:
            self.progress_var.set(0.0)
            messagebox.showerror("Erro", "A biblioteca reportlab não está instalada. Instale com: pip install reportlab")
        except Exception as e:
            self.progress_var.set(0.0)
            messagebox.showerror("Erro", str(e))



    def _config_vars(self) -> Dict[str, tk.Variable]:
        return {
            "thickness_mm": self.var_thickness,
            "cover_mm": self.var_cover,
            "concrete_class": self.var_concrete,
            "steel_fyk": self.var_steel,
            "local_y_is_vertical": self.var_y_vertical,
            "swap_axes": self.var_swap_axes,
            "limit_spacing_250": self.var_crack_spacing,
            "wk_check": self.var_crack_check,
            "qp_case": self.var_qp_case,
            "combo_type": self.var_combo_type,
            "wmax_mm": self.var_wmax,
            "moment_unit": self.var_moment_unit,
            "shear_unit": self.var_shear_unit,
            "phi_min": self.var_phi_min,
            "phi_max": self.var_phi_max,
            "optimize_rebar": self.var_optimize_rebar,
            "rebar_strategy": self.var_rebar_strategy,
            "base_vertical_phi": self.var_base_v_phi,
            "base_vertical_spacing": self.var_base_v_spacing,
            "base_horizontal_phi": self.var_base_h_phi,
            "base_horizontal_spacing": self.var_base_h_spacing,
            "mxy_method": self.var_wa_method,
            "reduce_governing_cases": self.var_reduce,
        }

    def save_config(self):
        path = filedialog.asksaveasfilename(
            title="Guardar configurações",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
        )
        if not path:
            return
        try:
            data = {k: v.get() for k, v in self._config_vars().items()}
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.status_var.set(f"Configurações guardadas: {path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def load_config(self):
        path = filedialog.askopenfilename(
            title="Carregar configurações",
            filetypes=[("JSON", "*.json"), ("Todos", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            vars_map = self._config_vars()
            for k, val in data.items():
                if k in vars_map:
                    vars_map[k].set(val)
            self.status_var.set(f"Configurações carregadas: {path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def export_csv(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
        )
        if not path:
            return
        try:
            self.df_results.to_csv(path, index=False, sep=";")
            self.status_var.set(f"Exportado: {path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))


if __name__ == "__main__":
    app = WallsEC2App()
    app.mainloop()
